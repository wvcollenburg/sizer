"""The admin catalog Export -> Import round-trip, which is used as a backup.

The pair had drifted apart: export_models() wrote its own sheet names ("CPU
Options") while _import_catalog_from_excel reads the template's ("Model CPU
Options"). Only the Models sheet overlapped, and _sheet_rows returns [] for a
missing sheet, so re-importing a backup produced models stripped of every CPU,
RAM, storage and NIC option — and reported success while doing it.

These lock down that a backup actually restores: sheet names line up, every CPU
spec/benchmark column survives, cost_tier and validated_only survive, catalog
entries no model references survive, and "replace" really replaces.

Run: .venv/bin/python -m pytest tests/test_catalog_backup.py -q
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

import pytest  # noqa: E402
from flask import Flask  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

from database import db  # noqa: E402
import orm_models as om  # noqa: E402
import auth_models  # noqa: F401,E402  - complete the mapper registry
import project_models  # noqa: F401,E402
import admin_routes as ar  # noqa: E402


@pytest.fixture()
def app():
    application = Flask(__name__)
    application.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///:memory:"
    application.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    db.init_app(application)
    with application.app_context():
        db.drop_all()
        db.create_all()
        _seed()
        yield application


def _seed():
    """One model wired to a CPU/NIC/drive, plus a CPU and a NIC that no model
    references — the entries a model-driven export would never reach."""
    cpu = om.CpuCatalog(
        description="Xeon Gold 6526Y 16C/32T 3.5GHz", cores=16, threads=32, ghz=3.5,
        make="Intel", family="Xeon Gold", generation="Emerald Rapids",
        model="6526Y", p_cores=16, e_cores=0, base_ghz=2.8,
        all_core_turbo_ghz=3.5, max_turbo_ghz=4.0,
        specrate_int=143.0, passmark_cpu_mark=48000, passmark_single=3100,
    )
    orphan_cpu = om.CpuCatalog(
        description="Xeon Silver 4510 12C/24T 2.4GHz", cores=12, threads=24, ghz=2.4,
        make="Intel", generation="Emerald Rapids", specrate_int=88.5,
    )
    nic = om.NicCatalog(description="25GbE SFP28 2-port (Intel E810)",
                        ports=2, speed="25GbE")
    orphan_nic = om.NicCatalog(description="100GbE QSFP28 2-port (Mellanox CX6)",
                               ports=2, speed="100GbE")
    drive = om.DriveCatalog(drive_type="NVMe", size_tb=3.84)
    db.session.add_all([cpu, orphan_cpu, nic, orphan_nic, drive])
    db.session.add(om.DriveTypeIops(drive_type="NVMe", iops=75000))
    db.session.add(om.SizingSetting(key="iops_read_fraction", value=0.70))
    db.session.flush()

    model = om.Model(name="HC5250D", status="Active", category="1U All-Flash",
                     form_factor="1U Rack", chassis="Dell R660", socket="dual",
                     psu="2x 800W", ram_slots=16, min_nodes=3,
                     cost_tier=27.5, validated_only=True, notes="test unit")
    db.session.add(model)
    db.session.flush()
    db.session.add(om.ModelCpuOption(model_id=model.id, cpu_id=cpu.id,
                                     quantity=2, sort_order=0))
    db.session.add(om.ModelNicOption(model_id=model.id, nic_id=nic.id,
                                     quantity=1, sort_order=0))
    db.session.add(om.RamOption(model_id=model.id, size_gb=256))
    sc = om.StorageConfig(model_id=model.id, storage_type="nvme_only",
                          drives_per_node=10)
    db.session.add(sc)
    db.session.flush()
    db.session.add(om.StorageConfigDrive(storage_config_id=sc.id, drive_id=drive.id))
    db.session.commit()


def _export(application):
    with application.test_request_context():
        resp = ar.export_models()
        resp.direct_passthrough = False
        data = resp.get_data()
    path = os.path.join(tempfile.gettempdir(), "sizer-catalog-test.xlsx")
    with open(path, "wb") as fh:
        fh.write(data)
    return path


def _wipe():
    db.drop_all()
    db.create_all()


def test_export_sheet_names_are_the_ones_the_importer_reads(app):
    """The drift that broke backups: every sheet the export writes must be one
    the importer actually looks for."""
    wb = load_workbook(_export(app))
    written = set(wb.sheetnames)
    wb.close()
    read = {"CPUs", "NICs", "Drives", "Models", "Model CPU Options",
            "Model RAM Options", "Model Storage", "Model Drive Options",
            "Model NIC Options", "Drive IOPS", "Sizing Settings"}
    assert written <= read, f"export writes sheets nobody reads: {written - read}"


def test_round_trip_restores_model_options(app):
    path = _export(app)
    _wipe()
    ar._import_catalog_from_excel(path)

    m = om.Model.query.filter_by(name="HC5250D").one()
    assert len(m.cpu_links) == 1 and m.cpu_links[0].quantity == 2
    assert m.cpu_links[0].cpu.description.startswith("Xeon Gold 6526Y")
    assert [r.size_gb for r in m.ram_options] == [256]
    assert len(m.nic_links) == 1 and m.nic_links[0].nic.speed == "25GbE"
    assert m.storage_config.drives_per_node == 10
    assert len(m.storage_config.drive_links) == 1
    assert m.storage_config.drive_links[0].drive.size_tb == 3.84


def test_round_trip_preserves_cost_and_validated_only(app):
    path = _export(app)
    _wipe()
    ar._import_catalog_from_excel(path)

    m = om.Model.query.filter_by(name="HC5250D").one()
    assert m.cost_tier == 27.5, "cost_tier feeds the ranker; a silent 5.0 changes sizing"
    assert m.validated_only is True


def test_round_trip_preserves_every_cpu_spec_column(app):
    path = _export(app)
    _wipe()
    ar._import_catalog_from_excel(path)

    cpu = om.CpuCatalog.query.filter_by(
        description="Xeon Gold 6526Y 16C/32T 3.5GHz").one()
    assert cpu.specrate_int == 143.0, "perf-based sizing reads specrate_int"
    assert cpu.passmark_cpu_mark == 48000
    assert cpu.passmark_single == 3100
    assert cpu.generation == "Emerald Rapids"
    assert cpu.make == "Intel" and cpu.family == "Xeon Gold"
    assert cpu.model == "6526Y"
    assert (cpu.p_cores, cpu.e_cores) == (16, 0)
    assert cpu.base_ghz == 2.8 and cpu.max_turbo_ghz == 4.0


def test_round_trip_keeps_catalog_entries_no_model_uses(app):
    path = _export(app)
    _wipe()
    ar._import_catalog_from_excel(path)

    assert om.CpuCatalog.query.filter_by(
        description="Xeon Silver 4510 12C/24T 2.4GHz").one().specrate_int == 88.5
    assert om.NicCatalog.query.filter_by(
        description="100GbE QSFP28 2-port (Mellanox CX6)").one().ports == 2


def test_round_trip_restores_sizing_config(app):
    path = _export(app)
    _wipe()
    ar._import_catalog_from_excel(path)

    assert om.DriveTypeIops.query.filter_by(drive_type="NVMe").one().iops == 75000
    assert om.SizingSetting.query.filter_by(key="iops_read_fraction").one().value == 0.70


def test_add_mode_leaves_existing_rows_alone(app):
    """Default mode must not rewrite the host's own edits."""
    path = _export(app)
    m = om.Model.query.filter_by(name="HC5250D").one()
    m.cost_tier = 99.0
    cpu = om.CpuCatalog.query.filter_by(
        description="Xeon Gold 6526Y 16C/32T 3.5GHz").one()
    cpu.specrate_int = 1.0
    db.session.commit()

    ar._import_catalog_from_excel(path, "add")

    assert om.Model.query.filter_by(name="HC5250D").one().cost_tier == 99.0
    assert om.CpuCatalog.query.filter_by(
        description="Xeon Gold 6526Y 16C/32T 3.5GHz").one().specrate_int == 1.0


def test_add_mode_fills_empty_spec_columns(app):
    """A restore onto a sparsely-seeded catalog should enrich, not skip."""
    path = _export(app)
    cpu = om.CpuCatalog.query.filter_by(
        description="Xeon Gold 6526Y 16C/32T 3.5GHz").one()
    cpu.specrate_int = None
    cpu.generation = None
    db.session.commit()

    ar._import_catalog_from_excel(path, "add")

    cpu = om.CpuCatalog.query.filter_by(
        description="Xeon Gold 6526Y 16C/32T 3.5GHz").one()
    assert cpu.specrate_int == 143.0
    assert cpu.generation == "Emerald Rapids"


def test_replace_mode_restores_over_local_edits(app):
    path = _export(app)
    m = om.Model.query.filter_by(name="HC5250D").one()
    m.cost_tier = 99.0
    cpu = om.CpuCatalog.query.filter_by(
        description="Xeon Gold 6526Y 16C/32T 3.5GHz").one()
    cpu.specrate_int = 1.0
    db.session.commit()

    ar._import_catalog_from_excel(path, "replace")

    m = om.Model.query.filter_by(name="HC5250D").one()
    assert m.cost_tier == 27.5
    assert len(m.cpu_links) == 1, "replace rebuilds the model's options"
    assert om.CpuCatalog.query.filter_by(
        description="Xeon Gold 6526Y 16C/32T 3.5GHz").one().specrate_int == 143.0
    assert om.Model.query.filter_by(name="HC5250D").count() == 1, "no duplicate"


def test_legacy_sheet_names_still_import(app):
    """Workbooks exported by builds before the rename must keep working."""
    _wipe()
    wb = Workbook()
    ws = wb.active
    ws.title = "Models"
    ws.append(["Name", "Status", "Category", "Form Factor", "Chassis", "Socket",
               "PSU", "RAM Slots", "Min Nodes", "Cost", "Validated Only", "Notes"])
    ws.append(["OLD-1", "Active", "1U", "1U Rack", "Chassis", "single",
               "2x 800W", 16, 3, 12.5, "No", None])
    ws_cpu = wb.create_sheet("CPU Options")      # old name
    ws_cpu.append(["Model Name", "Qty", "Description", "Cores", "Threads", "GHz"])
    ws_cpu.append(["OLD-1", 2, "Legacy Xeon 8C", 8, 16, 2.1])
    ws_ram = wb.create_sheet("RAM Options")      # old name
    ws_ram.append(["Model Name", "Size GB"])
    ws_ram.append(["OLD-1", 128])
    path = os.path.join(tempfile.gettempdir(), "sizer-catalog-legacy.xlsx")
    wb.save(path)

    ar._import_catalog_from_excel(path)

    m = om.Model.query.filter_by(name="OLD-1").one()
    assert len(m.cpu_links) == 1 and m.cpu_links[0].quantity == 2
    assert [r.size_gb for r in m.ram_options] == [128]
    assert m.cost_tier == 12.5


def test_import_reports_what_it_did(app):
    path = _export(app)
    _wipe()
    msg = ar._import_catalog_from_excel(path)["message"]
    assert "CPUs: 2 added" in msg
    assert "Models: 1 created" in msg


def test_template_and_export_share_one_schema(app):
    """The template is what people hand-author imports from; if its headers
    drift from the export's, a backup and a hand-built file stop being the same
    format — which is how the sheet names diverged in the first place."""
    from flask import Flask
    with Flask(__name__).test_request_context():
        resp = ar.catalog_template()
        resp.direct_passthrough = False
        data = resp.get_data()
    path = os.path.join(tempfile.gettempdir(), "sizer-catalog-template.xlsx")
    with open(path, "wb") as fh:
        fh.write(data)

    tpl = load_workbook(path)
    exp = load_workbook(_export(app))
    try:
        assert set(tpl.sheetnames) == set(exp.sheetnames)
        for name in tpl.sheetnames:
            assert [c.value for c in tpl[name][1]] == [c.value for c in exp[name][1]], \
                f"header drift on '{name}'"
    finally:
        tpl.close()
        exp.close()
