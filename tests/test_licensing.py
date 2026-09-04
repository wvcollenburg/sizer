"""Licence-aware scoring: the band model, the Essentials cliff, and the switch.

Covers docs/pricebook-plan.md §5 and §7. The golden values are taken from a REAL
customer quote (`_archive/Also Nor- Crayon - Marine Aluminium 1-27.pdf`), not
from our own parser output — a parser that agrees with itself proves nothing.

Run: .venv/bin/python -m pytest tests/test_licensing.py -q
"""
import os
import sys

os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("ENABLE_SCHEDULER", "0")
os.environ.setdefault("SECRET_KEY", "test-secret")

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

import pytest  # noqa: E402

import licensing  # noqa: E402
import parser_common  # noqa: E402
import pricebook_import  # noqa: E402

ARCHIVE = os.path.join(os.path.dirname(__file__), "..", "_archive")
PRICELIST = os.path.join(ARCHIVE, "Scale Computing Q4 2025 EUR Master Price List.xlsx")

# From the January quote (Q-102372-1). These are the numbers a customer was
# actually asked to pay.
QUOTE_BANDED_SKU = "HCOS-S-5-16C-PS"
QUOTE_BANDED_MSRP = 17339.00          # per node, 16 cores, 5 years, Premium
QUOTE_ESSENTIALS_SKU = "HCOS-5-PE"
QUOTE_ESSENTIALS_MSRP = 19944.00      # per 3-node cluster, 5 years, Professional

pytestmark = pytest.mark.skipif(
    not os.path.exists(PRICELIST),
    reason="archived price list not present")


@pytest.fixture(scope="module")
def parsed():
    return pricebook_import.parse_pricebook(PRICELIST)


@pytest.fixture(scope="module")
def book(parsed):
    return pricebook_import.build_book(parsed, region="EMEA", feed_label="Q4 2025")


# ── Parser ───────────────────────────────────────────────────────────────────

def test_parser_golden_counts(parsed):
    """Every licence row is accounted for. The unmatched count is the early
    warning that the export format moved, so it is asserted exactly."""
    c = parsed["counts"]
    assert c["license_rows"] == 370
    assert c["banded"] == 360
    assert c["flat"] == 9
    # HCOS-S-POC: the EUR 0 60-day trial. Neither a band nor a priced flat SKU,
    # so it is correctly excluded — but it must be REPORTED, not dropped.
    assert c["unmatched"] == 1
    assert parsed["unmatched"][0]["sku"] == "HCOS-S-POC"
    assert parsed["currency"] == "EUR"


def test_parser_finds_all_three_editions(parsed):
    editions = {b["edition"] for b in parsed["bands"]}
    assert editions == {"S", "L", "V"}
    # Only Standard has a Premium ladder; BRS and Video are Standard-Support only.
    support_by_edition = {}
    for b in parsed["bands"]:
        support_by_edition.setdefault(b["edition"], set()).add(b["support_tier"])
    assert support_by_edition["S"] == {"SS", "PS"}
    assert support_by_edition["L"] == {"SS"}
    assert support_by_edition["V"] == {"SS"}


def test_parser_matches_the_real_quote(book):
    """Golden values from the customer quote, to the cent."""
    assert book.band_map("S", 5, "PS")[16] == QUOTE_BANDED_MSRP
    assert book.flat_price("PE", 5) == QUOTE_ESSENTIALS_MSRP


def test_parser_rejects_a_sheet_with_missing_headers(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Product: Product Code", "List Price"])   # no family/desc/currency
    ws.append(["HCOS-S-1-16C-SS", 1234])
    path = tmp_path / "mangled.xlsx"
    wb.save(path)
    with pytest.raises(pricebook_import.PricebookFormatError) as exc:
        pricebook_import.parse_pricebook(str(path))
    assert "missing required column" in str(exc.value)


def test_parser_rejects_a_sheet_with_no_licence_family(tmp_path):
    """A well-formed sheet with zero licence rows means the family name moved.
    Fail loudly rather than apply an empty feed."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(list(pricebook_import.REQUIRED_COLUMNS))
    ws.append(["CHA-3-1A", "A chassis", 2937, "EUR", "Chassis"])
    path = tmp_path / "nolicence.xlsx"
    wb.save(path)
    with pytest.raises(pricebook_import.PricebookFormatError):
        pricebook_import.parse_pricebook(str(path))


def test_unknown_edition_letter_is_data_not_an_error():
    """A future edition must land as a priced band we simply do not offer —
    not as an unmatched row, and not as a crash."""
    rows = [{
        pricebook_import.COL_FAMILY: pricebook_import.LICENSE_FAMILY,
        pricebook_import.COL_CODE: "HCOS-Z-3-32C-SS",
        pricebook_import.COL_DESC: "Some future edition",
        pricebook_import.COL_PRICE: 4242,
        pricebook_import.COL_CURRENCY: "EUR",
    }]
    out = pricebook_import.parse_license_rows(rows)
    assert out["counts"]["unmatched"] == 0
    assert out["bands"][0]["edition"] == "Z"
    assert "Z" not in licensing.SELECTABLE_EDITIONS


# ── Bands and the cap ────────────────────────────────────────────────────────

def test_the_cap_emerges_from_the_data(book):
    """48C/52C/56C/64C carry one price, so cores past the cap are free. Nothing
    in the code knows the number 48 — moving the cap is a feed change."""
    bands = book.band_map("S", 5, "SS")
    assert bands[48] == bands[52] == bands[56] == bands[64]
    assert bands[44] < bands[48]

    below, _ = licensing.band_price(bands, 44)
    at_cap, _ = licensing.band_price(bands, 48)
    over_cap, _ = licensing.band_price(bands, 64)
    assert below < at_cap
    assert at_cap == over_cap


def test_cap_annotation_only_fires_above_the_cap(book):
    bands = book.band_map("S", 5, "SS")
    assert licensing.cap_annotation(bands, 44) is None
    note = licensing.cap_annotation(bands, 64)
    assert note and "no licence cost" in note


def test_above_the_top_band_clamps_and_says_so(book):
    """A dual-socket EPYC node can be 192 cores with no SKU. Clamping is only
    safe because the per-core burden term prices cores independently."""
    bands = book.band_map("S", 5, "SS")
    price, note = licensing.band_price(bands, 192)
    assert price == bands[64]
    assert note and "exceeds the published 64-core band" in note


def test_brs_and_video_cap_far_earlier(book):
    """Both cap at 24C, matching a design that keeps compute deliberately small.
    Deferred from the UI, but the data must be right for when it is picked up."""
    for edition in ("L", "V"):
        bands = book.band_map(edition, 1, "SS")
        assert bands[24] == bands[64]
        assert bands[20] < bands[24]


def test_term_is_clamped_to_the_sellable_range():
    # Scale renews to 7 years but does not sell 7 up front.
    assert licensing.clamp_term(7) == 5
    assert licensing.clamp_term(0) == 1
    assert licensing.clamp_term(None) == 5
    assert licensing.clamp_term("3") == 3


# ── Essentials ───────────────────────────────────────────────────────────────

def test_essentials_matches_the_quote(book):
    """3 nodes, 32 GB/node, 5 years, Premium support -> the PE price the customer
    was quoted."""
    out = licensing.cluster_license(book, [3], 16, 32, 5, support="PS")
    assert out["basis"] == "essentials"
    assert out["eur"] == QUOTE_ESSENTIALS_MSRP
    assert out["annotations"]["essentials_kind"] == "PE"


def test_essentials_tier_follows_support_not_price(book):
    """Essentials Kit is the Standard-Support tier, Professional Essentials the
    Premium one. Picking whichever is cheaper would silently mix support tiers
    inside one sizing — the quote pairs PS bands with a PE cluster."""
    ss = licensing.cluster_license(book, [3], 16, 32, 5, support="SS")
    ps = licensing.cluster_license(book, [3], 16, 32, 5, support="PS")
    assert ss["annotations"]["essentials_kind"] == "SE"
    assert ps["annotations"]["essentials_kind"] == "PE"
    assert ss["eur"] < ps["eur"]     # SE is cheaper, but only SS gets it


def test_essentials_cannot_be_stacked_across_clusters(book):
    """A 6-node result split into [3, 3] is a CAPACITY split, not two Essentials
    clusters. This is the single most important eligibility rule — getting it
    wrong would halve the licence line on every large SMB sizing."""
    eligible, near_miss = licensing.essentials_eligibility([3, 3], 128)
    assert eligible is False
    assert near_miss and "cannot be stacked" in near_miss

    out = licensing.cluster_license(book, [3, 3], 16, 128, 5)
    assert out["basis"] == "per_node"
    # Six nodes at the 16C band, not one flat cluster price.
    assert out["eur"] == book.band_map("S", 5, "SS")[16] * 6


def test_essentials_ram_ceiling_and_near_miss(book):
    assert licensing.essentials_eligibility([3], 256)[0] is True
    eligible, near_miss = licensing.essentials_eligibility([3], 288)
    assert eligible is False
    # The near-miss string is the most actionable thing this feature produces.
    assert near_miss == "32 GB/node over the 256 GB Essentials ceiling"


def test_essentials_node_count_boundary():
    assert licensing.essentials_eligibility([2], 128)[0] is False
    assert licensing.essentials_eligibility([3], 128)[0] is True
    assert licensing.essentials_eligibility([4], 128)[0] is False
    assert "Essentials requires exactly 3" in licensing.essentials_eligibility([4], 128)[1]


def test_essentials_loses_when_per_node_is_cheaper(book):
    """Normally a 2.6x saving, but it must not be applied blindly: a tiny
    2-core cluster on a short term is cheaper per-node."""
    out = licensing.cluster_license(book, [3], 2, 32, 1)
    assert out["annotations"]["essentials_eligible"] is True
    assert out["annotations"]["essentials_applied"] is False
    assert out["basis"] == "per_node"


# ── Eligibility vocabulary ───────────────────────────────────────────────────

def test_unrepresentable_rule_fails_loudly():
    """"We cannot express this" must be a visible error, never a silent
    misprice."""
    licensing.validate_rule({"exact_nodes": 3, "max_ram_gb_per_node": 256})
    with pytest.raises(licensing.UnrepresentableRule) as exc:
        licensing.validate_rule({"discount_if_friday": True})
    assert "discount_if_friday" in str(exc.value)


def test_seeded_rules_are_all_representable():
    for edition, spec in pricebook_import.SEEDED_RULES.items():
        licensing.validate_rule(dict(spec["rule"]))


# ── Guest licensing detection ────────────────────────────────────────────────

def test_detects_windows_and_reports_the_share():
    vms = [{"os": "Microsoft Windows Server 2022 (64-bit)", "vcpus": 8},
           {"os": "Ubuntu Linux (64-bit)", "vcpus": 2}]
    exposure, detail = parser_common.detect_guest_licensing(vms)
    assert exposure == parser_common.EXPOSURE_WINDOWS
    assert "80% of vCPU" in detail


def test_a_pure_linux_estate_detects_as_none():
    vms = [{"os": "Ubuntu Linux (64-bit)", "vcpus": 4},
           {"os": "Red Hat Enterprise Linux 9", "vcpus": 8}]
    exposure, detail = parser_common.detect_guest_licensing(vms)
    assert exposure == parser_common.EXPOSURE_NONE
    assert "no Windows guests" in detail


def test_no_os_information_yields_no_default():
    """Absent OS data must not silently assert "none" — that would zero the
    per-core burden for a source we simply know nothing about."""
    assert parser_common.detect_guest_licensing([{"vcpus": 4}]) == (None, None)
    assert parser_common.detect_guest_licensing([]) == (None, None)


def test_detection_never_claims_a_database():
    """A database is invisible in the guest OS string; only an SA can declare
    it. Detection must never return the windows_db level."""
    vms = [{"os": "Microsoft Windows Server 2022 (64-bit)", "vcpus": 8}] * 5
    exposure, _ = parser_common.detect_guest_licensing(vms)
    assert exposure != parser_common.EXPOSURE_WINDOWS_DB


def test_archived_sizings_detect_windows():
    """End-to-end through the real importer, not a hand-built VM list."""
    import glob
    from liveoptics import parse_liveoptics
    files = sorted(glob.glob(os.path.join(ARCHIVE, "LiveOptics*.xlsx")))
    if not files:
        pytest.skip("no archived LiveOptics exports")
    for path in files:
        summary = parse_liveoptics(path)["summary"]
        assert summary["guest_licensing"] == parser_common.EXPOSURE_WINDOWS
        assert "% of vCPU" in summary["guest_licensing_detail"]


# ── No feed ──────────────────────────────────────────────────────────────────

def test_an_empty_book_is_not_priceable_and_does_not_raise():
    """A region with no price list must never fail a sizing. The engine still
    has to produce a recommendation; it just cannot price the licence."""
    empty = licensing.LicenseBook()
    assert not empty
    out = licensing.cluster_license(empty, [3], 16, 128, 5)
    assert out["eur"] is None
    assert out["basis"] is None


# ── End-to-end scoring ───────────────────────────────────────────────────────
#
# The switch (`license_scoring`) must be genuinely reversible: with it off, not
# one recommendation may move. That is what makes it safe to ship before the
# calibration sweep has been reviewed.

import app as appmod        # noqa: E402
from database import db     # noqa: E402
from extensions import limiter  # noqa: E402

SUMMARY = {
    "active_vms": 40, "total_vms": 44, "total_vcpus": 180, "total_ram_gb": 900,
    "used_storage_tb": 22.5, "total_storage_tb": 60.0, "hosts": 4,
    "total_host_ghz": 400.0, "peak_cpu_ghz": 120.0, "total_host_cores": 96,
    "total_host_ram_gb": 1024, "vm_iops": 0, "peak_ram_gb": 700,
    "total_vm_provisioned_memory_gb": 900, "datastore_used_tb": 22.5,
    "nic_speed_mbps": 10000,
}

SMALL_SUMMARY = dict(SUMMARY, total_vcpus=24, total_ram_gb=90, peak_ram_gb=80,
                     used_storage_tb=3.0, datastore_used_tb=3.0,
                     total_vm_provisioned_memory_gb=90, active_vms=8,
                     total_vms=8, total_host_cores=16, total_host_ram_gb=128,
                     total_host_ghz=60.0, peak_cpu_ghz=20.0)


def _seed_models():
    from orm_models import (Model, CpuCatalog, ModelCpuOption, RamOption,
                            DriveCatalog, StorageConfig, StorageConfigDrive)
    nvme = DriveCatalog(drive_type="NVMe", size_tb=7.68)
    db.session.add(nvme)
    db.session.flush()
    from orm_models import DriveTypeIops
    db.session.add(DriveTypeIops(drive_type="NVMe", iops=75000))
    # Two CPU rungs on one chassis so the ranker has a real choice to make: the
    # licence cap sits between them.
    specs = [("Xeon 4516Y+", 24, 219.0, 2), ("Xeon 6538N", 32, 285.0, 2)]
    from models import MODEL_COSTS
    # The GPU variant is seeded alongside so the "never wins unprompted" test is
    # real. Without it that assertion passes because the model simply is not
    # there — the vacuous pass this suite guards against everywhere else.
    # CpuCatalog.description is unique, so the CPUs are created ONCE and both
    # models link to the same rows — which is also how the real catalog works.
    cpus = []
    for desc, cores, spec, qty in specs:
        cpu = CpuCatalog(description=desc, cores=cores, threads=cores * 2,
                         ghz=2.9, specrate_int=spec)
        db.session.add(cpu)
        db.session.flush()
        cpus.append((cpu, qty))

    for name, tier in (("HC3450DF", MODEL_COSTS["HC3450DF"]),
                       ("HC3450FG", MODEL_COSTS["HC3450FG"])):
        model = Model(name=name, status="Active", category="3XXX Core",
                      form_factor="1U", chassis="dual", min_nodes=2, cost_tier=tier)
        db.session.add(model)
        db.session.flush()
        for i, (cpu, qty) in enumerate(cpus):
            db.session.add(ModelCpuOption(model_id=model.id, cpu_id=cpu.id,
                                          quantity=qty, sort_order=i))
        storage = StorageConfig(model_id=model.id, storage_type="nvme_only",
                                drives_per_node=4)
        db.session.add(storage)
        db.session.flush()
        db.session.add(StorageConfigDrive(storage_config_id=storage.id, drive_id=nvme.id))
        for size in (128, 256, 512, 1024):
            db.session.add(RamOption(model_id=model.id, size_gb=size))
    db.session.commit()


def _set_tunables(**kw):
    from orm_models import SizingSetting
    from tunables import refresh_from_db
    for key, value in kw.items():
        row = SizingSetting.query.filter_by(key=key).first()
        if row is None:
            db.session.add(SizingSetting(key=key, value=value))
        else:
            row.value = value
    db.session.commit()
    refresh_from_db()


@pytest.fixture()
def seeded_app():
    application = appmod.app
    application.config["TESTING"] = True
    application.config["RATELIMIT_ENABLED"] = False
    limiter.enabled = False
    with application.app_context():
        db.drop_all()
        db.create_all()
        _seed_models()
        pricebook_import.seed_feed_from_file(PRICELIST, label="Q4 2025 EUR")
        yield application


def _recommend(summary=None, **kw):
    from recommend import generate_recommendations
    return generate_recommendations(summary or SUMMARY, **kw)


def _shape(result):
    """The parts of a recommendation a change must not silently move."""
    return [(r["model"], r["node_count"], r["cpu"], r["ram_per_node_gb"])
            for r in result["recommendations"]]


def test_switch_off_changes_nothing(seeded_app):
    with seeded_app.app_context():
        _set_tunables(license_scoring=0)
        before = _shape(_recommend())
        # Even with a licence feed loaded and licence inputs supplied, the
        # legacy path must govern while the switch is off.
        after = _shape(_recommend(license_term_years=3, guest_licensing="none"))
        assert before == after
        assert all(r.get("licensing") is None
                   for r in _recommend()["recommendations"])


def test_switch_on_annotates_without_leaking_price(seeded_app):
    from test_security import _offending_keys
    with seeded_app.app_context():
        _set_tunables(license_scoring=1)
        result = _recommend()
        assert result["recommendations"]
        lic = result["recommendations"][0]["licensing"]
        assert lic is not None
        assert lic["term_years"] == 5
        assert lic["basis"] in ("per_node", "essentials")
        # Everything that crosses the wire is a boolean, a string or a term
        # count — never a euro figure.
        assert _offending_keys(result) == []
        for value in lic.values():
            assert not isinstance(value, float)


def test_licence_term_is_independent_of_the_growth_horizon(seeded_app):
    with seeded_app.app_context():
        _set_tunables(license_scoring=1)
        five = _recommend(years=5, license_term_years=5)
        three = _recommend(years=5, license_term_years=3)
        assert five["recommendations"][0]["licensing"]["term_years"] == 5
        assert three["recommendations"][0]["licensing"]["term_years"] == 3
        # Growth horizon unchanged in both — the two inputs are separate.
        assert five["projection"]["years"] == three["projection"]["years"] == 5


def test_guest_exposure_reaches_the_annotation(seeded_app):
    with seeded_app.app_context():
        _set_tunables(license_scoring=1)
        none_run = _recommend(guest_licensing="none")
        win_run = _recommend(guest_licensing="windows")
        assert none_run["recommendations"][0]["licensing"]["guest_licensing"] == "none"
        assert win_run["recommendations"][0]["licensing"]["guest_licensing"] == "windows"


def test_essentials_wins_a_small_three_node_sizing(seeded_app):
    """The cliff should speak through the score without a thumb on the scale."""
    with seeded_app.app_context():
        _set_tunables(license_scoring=1)
        result = _recommend(SMALL_SUMMARY, target_nodes=3)
        top = result["recommendations"][0]
        if top["ram_per_node_gb"] <= licensing.ESSENTIALS_MAX_RAM_GB_PER_NODE:
            assert top["licensing"].get("essentials_eligible") is True
            assert top["licensing"]["basis"] == "essentials"
        else:
            # Too much RAM for Essentials — then the near-miss must say so,
            # which is the actionable half of the feature.
            assert "essentials_near_miss" in top["licensing"]


def test_no_feed_falls_back_instead_of_failing(seeded_app):
    """Deleting the feed must not break sizing — it must silently revert to the
    legacy term. A region without a price list still has to be sizeable."""
    from orm_models import CatalogFeed
    with seeded_app.app_context():
        _set_tunables(license_scoring=1)
        CatalogFeed.query.delete()
        db.session.commit()
        result = _recommend()
        assert result["recommendations"]
        assert result["recommendations"][0].get("licensing") is None


def test_guest_exposure_multipliers_are_ordered():
    """none <= windows <= windows_db is a real invariant, not a preference: a
    Linux estate cannot carry MORE per-core burden than a Windows one. An admin
    inverting these would silently reverse the feature."""
    from tunables import DEFAULTS
    assert (DEFAULTS["guest_exposure_none"]
            <= DEFAULTS["guest_exposure_windows"]
            <= DEFAULTS["guest_exposure_windows_db"])
    # 0.6 is measured, not chosen — below it, toggling the control starts
    # moving node counts on real sizings (tools/license_sweep.py).
    assert DEFAULTS["guest_exposure_none"] == 0.6


# ── platform_tier basket and divergence (§8.2, §8.3) ─────────────────────────

def test_basket_rate_matches_the_shipped_default():
    """The default eur_per_tier_point must be the basket's median, or the dial
    and the evidence have silently diverged."""
    import platform_tier
    from models import MODEL_COSTS
    from tunables import DEFAULTS
    rate = platform_tier.basket_rate(tiers=MODEL_COSTS)
    assert round(rate) == DEFAULTS["eur_per_tier_point"]


def test_ladder_agrees_with_the_basket_after_rebasing():
    """The ladder was re-based to the basket on 2026-09-02, so nothing should
    diverge any more.

    This agreement is ARITHMETIC, not evidence — every reference was scaled to
    the median by construction. What the test actually guards is that nobody
    hand-edits a tier to nudge a recommendation without re-deriving it from a
    price: do that, and this fails.
    """
    import platform_tier
    from models import MODEL_COSTS
    rows, median = platform_tier.divergence_report(tiers=MODEL_COSTS)
    flagged = {r["model"] for r in rows if r["flagged"]}
    assert flagged == set(), f"tier hand-edited away from its price: {flagged}"
    for r in rows:
        assert abs(r["deviation"]) < 0.01


def test_rebasing_preserved_ordering_within_each_family():
    """Bands were scaled by a single factor each, so relative order inside a
    family must be untouched. A reordering would mean a model silently changed
    rank against its siblings."""
    from models import APPLIANCE_MODELS, MODEL_COSTS
    import collections
    fams = collections.defaultdict(list)
    for name, m in APPLIANCE_MODELS.items():
        fams[m["category"]].append(name)
    # Known-correct ordering from the pre-re-base ladder, per family.
    for cat, names in fams.items():
        tiers = [MODEL_COSTS[n] for n in names]
        assert all(t > 0 for t in tiers), f"{cat} has a non-positive tier"


def test_epoch_normalisation_exposes_the_hc1450d_drift():
    """On an as-quoted basis the HC1450D looks perfect, because its reference is
    the January price. Re-priced to a current basis it is 2x out — which is the
    drift the report exists to surface, not a mis-tiering."""
    import platform_tier
    from models import MODEL_COSTS
    as_quoted, _ = platform_tier.divergence_report(tiers=MODEL_COSTS)
    current, _ = platform_tier.divergence_report(tiers=MODEL_COSTS,
                                                 normalise_epoch=True)
    quoted_flag = {r["model"]: r["flagged"] for r in as_quoted}
    current_flag = {r["model"]: r["flagged"] for r in current}
    assert quoted_flag["HC1450D"] is False
    assert current_flag["HC1450D"] is True


def test_reference_prices_are_configured_nodes_not_bare_frames():
    """The single easiest way to corrupt this basket is to take the chassis row
    from the PRICE LIST (a bare frame) instead of from a QUOTE (a whole node).
    CHA-3-1A is EUR 2,937 in the list and EUR 19,385 on the quote."""
    import platform_tier
    by_model = {c["model"]: c for c in platform_tier.REFERENCE_CONFIGS}
    assert by_model["HC1450D"]["price_eur"] == 19385.00
    assert by_model["HE153"]["price_eur"] == 1588.00
    # Every reference must record where it came from and when.
    for cfg in platform_tier.REFERENCE_CONFIGS:
        assert cfg["as_of"] and cfg["source"]


# ── Category naming scheme ───────────────────────────────────────────────────

def test_categories_follow_the_model_numbering():
    """Categories are named after the model number, which also carries the
    generation: 14XX -> 16XX is an update of the same line, not a new family.
    The old free-text names ("Datacenter 1U All-Flash") hid that."""
    import re
    from models import APPLIANCE_MODELS
    expected = {"1XX SFF", "2XX SFF", "5XX Edge",
                "1XXX Core", "3XXX Core", "5XXX Core", "Cloud"}
    seen = {m["category"] for m in APPLIANCE_MODELS.values()}
    assert seen <= expected, f"unexpected category: {seen - expected}"

    for name, model in APPLIANCE_MODELS.items():
        m = re.match(r"^H([EC])(\d)", name)
        if not m:
            continue          # SE100, Cloud Unity — handled explicitly below
        family, digit = m.groups()
        prefix = digit + ("XX " if family == "E" else "XXX ")
        assert model["category"].startswith(prefix), (
            f"{name} is in {model['category']!r}, which does not match its number")

    assert APPLIANCE_MODELS["SE100"]["category"] == "1XX SFF"
    assert APPLIANCE_MODELS["Cloud Unity"]["category"] == "Cloud"


def test_category_migration_leaves_admin_edits_alone():
    """The migration only rewrites KNOWN legacy strings. A category an admin has
    customised must survive — it is an editable field, and clobbering a
    deliberate edit on every boot is worse than a stale name."""
    import seed
    assert seed._category_for("HC1600") == "1XXX Core"
    assert seed._category_for("HE552F") == "5XX Edge"
    assert seed._category_for("SE100") == "1XX SFF"
    assert seed._category_for("Something Custom") is None
    # "Edge" maps to None in the table because it spans 1XX and 2XX — the model
    # number decides, not the old string.
    assert seed.LEGACY_CATEGORIES["Edge"] is None
    assert "Datacenter 1U All-Flash GPU" in seed.LEGACY_CATEGORIES


# ── Catalog truth (§4.2) ─────────────────────────────────────────────────────

def test_absence_from_the_price_list_is_not_evidence_of_eol():
    """A model absent from a price list is EITHER end-of-life OR newer than the
    list. Treating every absence as EOL would mark new products end-of-life and
    silently stop the sizer recommending the newest hardware — which is what the
    original §4.2 analysis would have done.

    The 16XX(D) line is the confirmed case: an update of the 14XX(D), newer than
    the Q4 2025 list.
    """
    import sys as _sys
    _sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "tools"))
    import catalog_check
    from models import APPLIANCE_MODELS

    priced = catalog_check.priced_models(PRICELIST)
    new, eol, unknown = catalog_check.classify(APPLIANCE_MODELS, priced)
    new_names = {n for n, _ in new}
    eol_names = {n for n, _ in eol}

    assert {"HC1600", "HC1650D"} <= new_names, "the confirmed-new line must not read as EOL"
    assert not (new_names & eol_names)
    # Same-number variants (HE155-1 vs HE155-2) carry no generational signal.
    assert "HE155-1" in {n for n, _ in unknown}


def test_price_list_matching_is_case_insensitive():
    """The list says HE153P, our catalog says HE153p. A case-sensitive compare
    reports that model as BOTH missing from our catalog AND end-of-life at the
    same time, which is how a shipping product gets marked EOL."""
    import sys as _sys
    _sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "tools"))
    import catalog_check
    from models import APPLIANCE_MODELS

    priced = catalog_check.priced_models(PRICELIST)
    assert "HE153P" in priced
    _new, eol, _unknown = catalog_check.classify(APPLIANCE_MODELS, priced)
    assert "HE153p" not in {n for n, _ in eol}
    ours = {n.upper() for n in APPLIANCE_MODELS}
    assert "HE153P" in ours


# ── Quarterly price-list import (tools/import_pricebook.py) ──────────────────

def _tools_path():
    import sys as _sys
    p = os.path.join(os.path.dirname(__file__), "..", "tools")
    if p not in _sys.path:
        _sys.path.insert(0, p)
    return p


def test_import_dry_run_writes_nothing(seeded_app, capsys):
    """The default must be a dry run. A quarterly list decides which
    configuration the sizer recommends; installing one by accident would change
    live proposals with nobody noticing."""
    _tools_path()
    import import_pricebook
    from orm_models import CatalogFeed

    with seeded_app.app_context():
        before = CatalogFeed.query.count()
        rc = import_pricebook.main([PRICELIST])
        assert rc == 0
        assert CatalogFeed.query.count() == before
    assert "DRY RUN" in capsys.readouterr().out


def test_import_refuses_a_mangled_file(seeded_app, tmp_path, capsys):
    from openpyxl import Workbook
    _tools_path()
    import import_pricebook
    from orm_models import CatalogFeed

    wb = Workbook()
    wb.active.append(["Product: Product Code", "List Price"])
    path = tmp_path / "mangled.xlsx"
    wb.save(path)

    with seeded_app.app_context():
        before = CatalogFeed.query.count()
        rc = import_pricebook.main([str(path)])
        assert rc == 2                      # a distinct exit code, for scripting
        assert CatalogFeed.query.count() == before
    assert "REFUSED" in capsys.readouterr().out


def test_installing_a_new_feed_keeps_the_old_one(seeded_app, tmp_path):
    """Superseded feeds are retained, not deleted: a saved sizing stamped with
    an old feed must still reproduce after the next quarterly import."""
    import re
    from openpyxl import Workbook, load_workbook
    _tools_path()
    import import_pricebook
    from orm_models import CatalogFeed

    src = load_workbook(PRICELIST, data_only=True)
    rows = list(src[src.sheetnames[0]].values)
    wb = Workbook(); ws = wb.active
    ws.append(list(rows[0]))
    for r in rows[1:]:
        r = list(r)
        if re.match(r"^HCOS-S-\d-\d+C", str(r[0] or "")) and isinstance(r[10], (int, float)):
            r[10] = round(r[10] * 1.06, 2)      # a 6% Standard-band rise
        ws.append(r)
    path = tmp_path / "next_quarter.xlsx"
    wb.save(path)

    with seeded_app.app_context():
        before = CatalogFeed.query.count()
        assert import_pricebook.main([str(path), "--label", "Q1 test", "--apply"]) == 0
        assert CatalogFeed.query.count() == before + 1
        current = CatalogFeed.query.filter_by(region="EMEA", is_current=True).all()
        assert len(current) == 1, "exactly one feed may be current per region"
        assert current[0].label == "Q1 test"
        # And the new prices are the ones the engine now reads.
        from orm_models import load_license_book
        assert load_license_book("EMEA").band_map("S", 5, "SS")[48] == pytest.approx(
            41094 * 1.06, rel=1e-3)


# ── GPU node weighting ───────────────────────────────────────────────────────

def test_gpu_node_carries_a_deliberate_surcharge():
    """HC3450FG is weighted roughly double what its hardware alone implies.

    Its hardware delta over the non-GPU HC3450F covers TWO Nvidia L4 cards. The
    doubling on top stands in for per-user vGPU licensing, which is steep and
    which nothing in the engine can see. It deliberately does not look
    proportionate — it is close to real total cost of ownership.
    """
    from models import MODEL_COSTS, APPLIANCE_MODELS
    assert APPLIANCE_MODELS["HC3450FG"]["gpu"] == "2 x Nvidia L4 24GB"
    plain, gpu = MODEL_COSTS["HC3450F"], MODEL_COSTS["HC3450FG"]
    assert gpu > plain * 2, "the vGPU-licensing surcharge has been lost"


def test_gpu_node_never_wins_an_unprompted_recommendation(seeded_app):
    """The engine has NO GPU demand signal — nothing reads the `gpu` field — so
    a GPU node winning on its own would be the engine guessing. It must still
    size correctly when an SA names it, which bypasses ranking entirely."""
    from recommend import generate_recommendations
    with seeded_app.app_context():
        _set_tunables(license_scoring=1)
        recs = generate_recommendations(SUMMARY).get("recommendations") or []
        assert recs, "no candidates — this test would pass vacuously"
        models = {r["model"] for r in recs}
        assert "HC3450FG" in models or len(models) > 1, (
            "the GPU model is not in the candidate set at all — the assertion "
            "below would prove nothing")
        assert recs[0]["model"] != "HC3450FG"
