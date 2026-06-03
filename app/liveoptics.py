from openpyxl import load_workbook


def parse_liveoptics(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    # Live Optics ships two unrelated workbook shapes under the same brand:
    #   * the VMware/vSphere collector ("ESX Hosts" + "VMs" + perf sheets), and
    #   * the GENERAL server collector (a single "Servers" sheet, one row per
    #     OS instance) used for Hyper-V and physical/mixed estates.
    # The GENERAL export has no host/guest split and folds performance into the
    # inventory row, so it gets its own parser. Both paths return the same dict
    # shape so everything downstream (summary, recommend, UI) is format-blind.
    if "Servers" in wb.sheetnames:
        result = _parse_general(wb)
    else:
        result = {
            "project": _parse_details(wb),
            "hosts": _parse_hosts(wb),
            "host_performance": _parse_host_perf(wb),
            "datastores": _parse_datastores(wb),
            "vms": _parse_vms(wb),
            "vm_performance": _parse_vm_perf(wb),
            "host_disks": _parse_host_disks(wb),
            "host_nics": _parse_host_nics(wb),
        }
    wb.close()

    result["summary"] = _build_summary(result)
    if result.get("scan_type") == "general":
        _finalize_general_summary(result)
    return result


def _sheet_rows(wb, name):
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:]]


def _parse_details(wb):
    if "Details" not in wb.sheetnames:
        return {}
    ws = wb["Details"]
    info = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] and row[1]:
            info[str(row[0]).strip()] = row[1]
    return info


def _parse_general(wb):
    """Parse the GENERAL (Hyper-V / physical / mixed) Live Optics export.

    The GENERAL collector emits one "Servers" row per OS instance with no
    hypervisor host/guest split and performance folded into the same row. For
    sizing a consolidation onto HyperCore, every server *is* a workload, so each
    one becomes both:
      * a VM (the demand we pack onto the new cluster), and
      * a host (the current-estate footprint shown in the summary, and the basis
        for the GHz/core figures the recommender sizes against).
    With no overcommit data the natural vCPU:core ratio is 1:1 (the user can
    lower it in the UI). All server-local storage must land on the new shared
    pool, so it's recorded as cluster (primary) capacity, not local.
    """
    hosts, vms, perfs, datastores, disks = [], [], [], [], []

    # Per-server disk totals, used as a fallback when a Servers row omits its
    # roll-up capacity (some collectors leave it blank and only fill Server Disks).
    disk_totals = {}
    for r in _sheet_rows(wb, "Server Disks"):
        host = str(r.get("Server Name", "")).strip()
        agg = disk_totals.setdefault(host, {"cap": 0.0, "used": 0.0, "free": 0.0})
        agg["cap"] += _float(r.get("Capacity (GiB)", 0))
        agg["used"] += _float(r.get("Used Capacity (GiB)", 0))
        agg["free"] += _float(r.get("Free Capacity (GiB)", 0))

    for r in _sheet_rows(wb, "Servers"):
        name = str(r.get("Server Name", "")).strip()
        cores = _int(r.get("CPU Cores", 0))
        sockets = _int(r.get("CPU Sockets", 0))
        clock = _float(r.get("CPU Clock Speed (GHz)", 0))
        net_ghz = _float(r.get("Net Clock Speed (GHz)", 0))
        mem_kib = _float(r.get("Memory (KiB)", 0))
        peak_mem_kib = _float(r.get("Peak Memory Usage (KiB)", 0))
        mem_gb = round(mem_kib / 1048576, 2)
        peak_mem_gb = round(peak_mem_kib / 1048576, 2)
        local_cap = _float(r.get("Local Capacity (GiB)", 0))
        used_cap = _float(r.get("Used Capacity (GiB)", 0))
        free_cap = _float(r.get("Free Capacity (GiB)", 0))
        if local_cap == 0 and used_cap == 0:
            agg = disk_totals.get(name)
            if agg:
                local_cap, used_cap, free_cap = agg["cap"], agg["used"], agg["free"]
        os_name = r.get("OS", "")

        hosts.append({
            "name": name,
            "cluster": "",
            "manufacturer": r.get("Manufacturer", ""),
            "model": r.get("Model", ""),
            "cpu_sockets": sockets,
            "cpu_cores": cores,
            # No SMT/thread count in this export; cores is the only honest value.
            "cpu_threads": cores,
            "cpu_desc": r.get("CPU Description", ""),
            "cpu_ghz": clock,
            "net_ghz": net_ghz,
            "memory_kib": mem_kib,
            "memory_gb": round(mem_kib / 1048576, 1),
            "local_capacity_gib": local_cap,
            "vm_count": 1,
            "nic_count": 0,
        })

        vms.append({
            "name": name,
            "powered_on": True,        # only running/measured servers are collected
            "is_template": False,
            "os": os_name,
            "model": f"{r.get('Manufacturer', '')} {r.get('Model', '')}".strip(),
            "vcpus": cores,            # physical cores -> vCPU demand, 1:1
            "provisioned_memory_gb": mem_gb,
            "used_memory_gb": peak_mem_gb,
            "consumed_memory_gb": peak_mem_gb,
            "disk_capacity_gb": local_cap,
            "disk_used_gb": used_cap,
            "vdisk_size_gb": local_cap,
            "vdisk_used_gb": used_cap,
            "datastore": "",
            "host": name,
            "cluster": "",
        })

        peak_cpu_pct = _float(r.get("Peak CPU Percentage", 0))
        perfs.append({
            "host": name,
            "peak_cpu_pct": peak_cpu_pct,
            "peak_cpu_ghz": round(peak_cpu_pct / 100 * net_ghz, 2),
            # GENERAL export reports only peak CPU/memory, no averages.
            "avg_cpu_pct": 0,
            "avg_cpu_ghz": 0,
            "peak_mem_pct": round(peak_mem_kib / mem_kib * 100, 1) if mem_kib else 0,
            "peak_mem_mib": round(peak_mem_kib / 1024, 1),
            "avg_mem_pct": 0,
            "avg_mem_mib": 0,
            "peak_iops": _float(r.get("Peak IOPS", 0)),
            "avg_iops": _float(r.get("Average IOPS", 0)),
            "p95_iops": _float(r.get("95% IOPS", 0)),
            "peak_throughput_mbs": _float(r.get("Peak Throughput MB/s", 0)),
            "avg_throughput_mbs": _float(r.get("Avg. Throughput MB/s", 0)),
        })

        # Server-local storage must all move to the new shared pool, so it's the
        # primary ("cluster") sizing basis rather than excludable local capacity.
        datastores.append({
            "name": name,
            "type": "cluster",
            "capacity_gib": local_cap,
            "used_gib": used_cap,
            "free_gib": free_cap,
            "vm_count": 1,
        })

    for r in _sheet_rows(wb, "Server Disks"):
        disks.append({
            "host": str(r.get("Server Name", "")).strip(),
            "disk_name": r.get("Device Name", ""),
            "capacity_mib": _float(r.get("Capacity (GiB)", 0)) * 1024,
            "model": r.get("Model", ""),
            "vendor": r.get("Vendor", ""),
            "is_ssd": False,
        })

    return {
        "project": _parse_details(wb),
        "scan_type": "general",
        "hosts": hosts,
        "host_performance": perfs,
        "datastores": datastores,
        "vms": vms,
        "vm_performance": [],
        "host_disks": disks,
        "host_nics": [],
    }


def _finalize_general_summary(result):
    """Adjust summary fields that don't fit a mixed server estate."""
    summary = result["summary"]
    n = summary.get("host_count", 0)
    summary["current_platform"] = f"Mixed estate ({n} servers)"
    # hosts and VMs are the same physical boxes here, so the host-count tile is
    # redundant with the VM count; make the redundancy explicit and honest.
    summary["scan_type"] = "general"
    # A server-level scan exposes no overcommit, so the measured vCPU:core ratio
    # is a meaningless 1:1. Fall back to the standard 3:1 consolidation default
    # and flag it as assumed so the UI labels it as such (and doesn't recompute
    # it back to 1:1 when VMs are excluded).
    summary["vcpu_per_core_ratio"] = 3.0
    summary["vcpu_ratio_assumed"] = True


def _parse_hosts(wb):
    hosts = []
    for r in _sheet_rows(wb, "ESX Hosts"):
        hosts.append({
            "name": r.get("Host Name", ""),
            "cluster": r.get("Cluster", ""),
            "manufacturer": r.get("Manufacturer", ""),
            "model": r.get("Model", ""),
            "cpu_sockets": _int(r.get("CPU Sockets", 0)),
            "cpu_cores": _int(r.get("CPU Cores", 0)),
            "cpu_threads": _int(r.get("CPU Threads", 0)),
            "cpu_desc": r.get("CPU Description", ""),
            "cpu_ghz": _float(r.get("CPU Clock Speed (GHz)", 0)),
            "net_ghz": _float(r.get("Net Clock Speed (GHz)", 0)),
            "memory_kib": _float(r.get("Memory (KiB)", 0)),
            "memory_gb": round(_float(r.get("Memory (KiB)", 0)) / 1048576, 1),
            "local_capacity_gib": _float(r.get("Local Capacity (GiB)", 0)),
            "vm_count": _int(r.get("Guest VM Count", 0)),
            "nic_count": _int(r.get("Number of NICs", 0)),
        })
    return hosts


def _parse_host_perf(wb):
    perfs = []
    for r in _sheet_rows(wb, "ESX Performance"):
        perfs.append({
            "host": r.get("Host", ""),
            "peak_cpu_pct": _float(r.get("Peak CPU %", 0)),
            "peak_cpu_ghz": _float(r.get("Peak CPU (GHz)", 0)),
            "avg_cpu_pct": _float(r.get("Average CPU %", 0)),
            "avg_cpu_ghz": _float(r.get("Average CPU (GHz)", 0)),
            "peak_mem_pct": _float(r.get("Peak Memory %", 0)),
            "peak_mem_mib": _float(r.get("Peak Memory (MiB)", 0)),
            "avg_mem_pct": _float(r.get("Average Memory %", 0)),
            "avg_mem_mib": _float(r.get("Average Memory (MiB)", 0)),
            "peak_iops": _float(r.get("Peak IOPS", 0)),
            "avg_iops": _float(r.get("Average IOPS", 0)),
            "p95_iops": _float(r.get("95% IOPS", 0)),
            "peak_throughput_mbs": _float(r.get("Peak Throughput MB/s", 0)),
            "avg_throughput_mbs": _float(r.get("Avg Throughput MB/s", 0)),
        })
    return perfs


def _parse_datastores(wb):
    stores = []
    seen_cluster = set()
    for r in _sheet_rows(wb, "Host Devices"):
        dtype = str(r.get("Device Type", "")).strip()
        if dtype not in ("Cluster", "Local"):
            continue
        cap = _float(r.get("Capacity (GiB)", 0))
        used = _float(r.get("Used Capacity (GiB)", 0))
        free = _float(r.get("Free Capacity (GiB)", 0))
        vmc = _int(r.get("VM Count", 0))
        if dtype == "Cluster":
            # A shared datastore is listed once per host that mounts it, each
            # under a host-local name ("Disk 2" on one host == "Disk 7" on
            # another). Deduping on the device name therefore never fires and
            # the datastore is counted once per host. Dedup on a host-independent
            # signature instead so each LUN counts once (matches the LO
            # dashboard's capacity figures exactly).
            sig = (round(cap), round(used), round(free), vmc)
            if sig in seen_cluster:
                continue
            seen_cluster.add(sig)
        stores.append({
            "name": str(r.get("Device Name", "")).strip(),
            "type": "local" if dtype == "Local" else "cluster",
            "capacity_gib": cap,
            "used_gib": used,
            "free_gib": free,
            "vm_count": vmc,
        })
    return stores


def _parse_vms(wb):
    vms = []
    for r in _sheet_rows(wb, "VMs"):
        powered_on = str(r.get("Power State", "")).lower() == "poweredon"
        is_template = str(r.get("Template", "")).upper() == "TRUE"
        prov_mem_mib = _float(r.get("Provisioned Memory (MiB)", 0))
        used_mem_mib = _float(r.get("Used Memory (active) (MiB)", 0))
        consumed_mem_mib = _float(r.get("Consumed Memory (MiB)", 0))
        disk_cap_mib = _float(r.get("Guest VM Disk Capacity (MiB)", 0))
        disk_used_mib = _float(r.get("Guest VM Disk Used (MiB)", 0))
        vdisk_size_mib = _float(r.get("Virtual Disk Size (MiB)", 0))
        vdisk_used_mib = _float(r.get("Virtual Disk Used (MiB)", 0))

        vms.append({
            "name": r.get("VM Name", ""),
            "powered_on": powered_on,
            "is_template": is_template,
            "os": r.get("VM OS", ""),
            "vcpus": _int(r.get("Virtual CPU", 0)),
            "provisioned_memory_gb": round(prov_mem_mib / 1024, 2),
            "used_memory_gb": round(used_mem_mib / 1024, 2),
            "consumed_memory_gb": round(consumed_mem_mib / 1024, 2),
            "disk_capacity_gb": round(disk_cap_mib / 1024, 2),
            "disk_used_gb": round(disk_used_mib / 1024, 2),
            "vdisk_size_gb": round(vdisk_size_mib / 1024, 2),
            "vdisk_used_gb": round(vdisk_used_mib / 1024, 2),
            "datastore": r.get("Datastore", ""),
            "host": r.get("Host", ""),
            "cluster": r.get("Cluster", ""),
        })
    return vms


def _parse_vm_perf(wb):
    perfs = []
    for r in _sheet_rows(wb, "VM Performance"):
        perfs.append({
            "name": r.get("VM Name", ""),
            "peak_vcpu_pct": _float(r.get("Peak vCPU %", 0)),
            "peak_vcpu_ghz": _float(r.get("Peak vCPU (GHz)", 0)),
            "avg_vcpu_pct": _float(r.get("Average vCPU %", 0)),
            "avg_vcpu_ghz": _float(r.get("Average vCPU (GHz)", 0)),
            "peak_mem_pct": _float(r.get("Peak Memory %", 0)),
            "peak_mem_mib": _float(r.get("Peak Memory (MiB)", 0)),
            "avg_mem_mib": _float(r.get("Avg Memory (MiB)", 0)),
            "peak_iops": _float(r.get("Peak IOPS", 0)),
            "avg_iops": _float(r.get("Average IOPS", 0)),
        })
    return perfs


def _parse_host_disks(wb):
    disks = []
    for r in _sheet_rows(wb, "Host Disks"):
        disks.append({
            "host": r.get("Host", ""),
            "disk_name": r.get("Disk Name", ""),
            "capacity_mib": _float(r.get("Disk Capacity (MiB)", 0)),
            "model": r.get("Disk Model", ""),
            "vendor": r.get("Disk Vendor", ""),
            "is_ssd": str(r.get("SSD", "")).upper() == "TRUE",
        })
    return disks


def _parse_host_nics(wb):
    nics = []
    seen_speeds = set()
    for r in _sheet_rows(wb, "Host Network Adapters"):
        speed = _float(r.get("PNIC Speed (Mb/sec)", 0))
        seen_speeds.add(speed)
        nics.append({
            "host": r.get("Host", ""),
            "name": r.get("PNIC Name", ""),
            "speed_mbps": speed,
            "vendor": r.get("PNIC Vendor", ""),
            "device": r.get("PNIC Device Name", ""),
        })
    return nics


def _build_summary(data):
    hosts = data["hosts"]
    vms = data["vms"]
    perfs = data["host_performance"]
    datastores = data["datastores"]

    active_vms = [v for v in vms if v["powered_on"] and not v["is_template"]]

    total_host_cores = sum(h["cpu_cores"] for h in hosts)
    total_host_threads = sum(h["cpu_threads"] for h in hosts)
    total_host_ghz = sum(h["cpu_ghz"] * h["cpu_cores"] for h in hosts)
    total_host_ram_gb = sum(h["memory_gb"] for h in hosts)

    total_vcpus = sum(v["vcpus"] for v in active_vms)
    total_vm_prov_mem_gb = sum(v["provisioned_memory_gb"] for v in active_vms)
    total_vm_used_mem_gb = sum(v["consumed_memory_gb"] for v in active_vms)
    total_vm_disk_prov_gb = sum(v["vdisk_size_gb"] for v in active_vms)
    total_vm_disk_used_gb = sum(v["vdisk_used_gb"] for v in active_vms)

    # Cluster (shared) storage is the default sizing basis; local (per-host)
    # storage is added only when the user opts in (it holds ISOs/templates that
    # land on the cluster under HyperCore).
    cluster_total_gib = sum(d["capacity_gib"] for d in datastores if d.get("type") != "local")
    cluster_used_gib = sum(d["used_gib"] for d in datastores if d.get("type") != "local")
    local_total_gib = sum(d["capacity_gib"] for d in datastores if d.get("type") == "local")
    local_used_gib = sum(d["used_gib"] for d in datastores if d.get("type") == "local")

    peak_cpu_pct = max((p["peak_cpu_pct"] for p in perfs), default=0)
    avg_cpu_pct = sum(p["avg_cpu_pct"] for p in perfs) / len(perfs) if perfs else 0
    peak_cpu_ghz = sum(p["peak_cpu_ghz"] for p in perfs)
    avg_cpu_ghz = sum(p["avg_cpu_ghz"] for p in perfs)
    peak_mem_pct = max((p["peak_mem_pct"] for p in perfs), default=0)
    avg_mem_pct = sum(p["avg_mem_pct"] for p in perfs) / len(perfs) if perfs else 0
    total_peak_iops = sum(p["peak_iops"] for p in perfs)
    total_avg_iops = sum(p["avg_iops"] for p in perfs)
    total_p95_iops = sum(p.get("p95_iops", 0) for p in perfs)

    nic_speeds = set()
    for n in data.get("host_nics", []):
        if n["speed_mbps"] > 0:
            nic_speeds.add(n["speed_mbps"])

    return {
        "host_count": len(hosts),
        "cluster_name": hosts[0]["cluster"] if hosts else "",
        "current_platform": f"{hosts[0]['manufacturer']} {hosts[0]['model']}" if hosts else "",

        "total_host_cores": total_host_cores,
        "total_host_threads": total_host_threads,
        "total_host_ghz": round(total_host_ghz, 1),
        "total_host_ram_gb": round(total_host_ram_gb, 1),
        "per_host_cores": round(total_host_cores / len(hosts), 1) if hosts else 0,
        "per_host_ram_gb": round(total_host_ram_gb / len(hosts), 1) if hosts else 0,

        "total_vms": len(vms),
        "active_vms": len(active_vms),
        "total_vcpus": total_vcpus,
        "total_vm_provisioned_memory_gb": round(total_vm_prov_mem_gb, 1),
        "total_vm_used_memory_gb": round(total_vm_used_mem_gb, 1),
        "total_vm_provisioned_storage_gb": round(total_vm_disk_prov_gb, 1),
        "total_vm_used_storage_gb": round(total_vm_disk_used_gb, 1),
        "total_vm_provisioned_storage_tb": round(total_vm_disk_prov_gb / 1024, 2),
        "total_vm_used_storage_tb": round(total_vm_disk_used_gb / 1024, 2),

        "datastore_total_tb": round(cluster_total_gib / 1024, 2),
        "datastore_used_tb": round(cluster_used_gib / 1024, 2),
        "local_total_tb": round(local_total_gib / 1024, 2),
        "local_used_tb": round(local_used_gib / 1024, 2),
        "local_used_gb": round(local_used_gib),

        "peak_cpu_pct": round(peak_cpu_pct, 1),
        "avg_cpu_pct": round(avg_cpu_pct, 1),
        "peak_cpu_ghz": round(peak_cpu_ghz, 1),
        "avg_cpu_ghz": round(avg_cpu_ghz, 1),
        "peak_mem_pct": round(peak_mem_pct, 1),
        "avg_mem_pct": round(avg_mem_pct, 1),
        "total_peak_iops": round(total_peak_iops),
        "total_avg_iops": round(total_avg_iops),
        "p95_iops": round(total_p95_iops),

        "nic_speed_mbps": max(nic_speeds) if nic_speeds else 0,

        "vcpu_per_core_ratio": round(total_vcpus / total_host_cores, 2) if total_host_cores > 0 else 0,
        "vcpu_per_thread_ratio": round(total_vcpus / total_host_threads, 2) if total_host_threads > 0 else 0,

        "max_vm_ram_gb": max((v["provisioned_memory_gb"] for v in active_vms), default=0),
        "max_vm_cores": max((v["vcpus"] for v in active_vms), default=0),
    }


def _float(v):
    try:
        return float(v) if v else 0.0
    except (ValueError, TypeError):
        return 0.0


def _int(v):
    try:
        return int(v) if v else 0
    except (ValueError, TypeError):
        return 0
