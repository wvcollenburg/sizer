import io
import os
import secrets
import tempfile
from urllib.parse import urlparse

from flask import Flask, render_template, jsonify, request, send_file
from werkzeug.middleware.proxy_fix import ProxyFix
from database import init_db
from extensions import limiter
from auth import register_auth, start_scheduler, current_user
from projects import register_projects
from sqlalchemy.orm import joinedload
from orm_models import (
    Model, StorageConfig,
    ModelCpuOption, ModelNicOption, StorageConfigDrive,
    ValidatedNic, ValidatedPlatform,
    DriveCatalog, RamOption,
)
from models import RAM_SIZES_GB
from liveoptics import parse_liveoptics
from rvtools import parse_rvtools
from import_checks import build_import_warnings
from recommend import generate_recommendations
from tunables import T, refresh_from_db
from export_pptx import generate_proposal, generate_config_slide, generate_bundle_proposal
from export_docx import (build_proposal_docx, build_bundle_proposal_docx,
                         convert_docx_to_pdf, convert_pptx_to_pdf)
from export_gate import export_gate
from admin_routes import admin_bp
from i18n import SUPPORTED_LANGS, LANG_NAMES
from calc import (calculate_appliance, calculate_validated, MAX_STORAGE_ONLY_COUNT,
                  SNS_NO_REDUNDANCY_MSG)
from fingerprint import PARSER_VERSION

# Upper bounds on client-supplied sizing counts. The recommendation engine is
# internally bounded (it searches a small node range), but the direct
# /api/calculate path builds per-node structures + an SVG straight from these, so
# an unclamped value (e.g. node_count=1e8) is a trivial CPU/memory DoS. These
# ceilings are far above any real cluster while keeping a hostile request cheap.
MAX_NODE_COUNT = 1000
# A bundle export iterates every section into one document; bound it so a giant
# sections list can't drive an unbounded (and soffice-backed) render. A "section"
# is one source cluster today; with projects it is also one sizing's cluster
# (docs/projects-plan.md §7.1), which is why the limit is no longer named after
# clusters.
MAX_EXPORT_SECTIONS = 50


def _validated_disk_sizes():
    """Disk-size options for the validated picker, read live from the
    admin-editable drive catalog so a newly added drive size is immediately
    selectable without a code change. Keyed by performance bucket; the front end
    maps the spinning interface types (SAS/NLSAS/SATA) onto the HDD bucket."""
    buckets = {"HDD": set(), "SSD": set(), "NVMe": set()}
    for drive in DriveCatalog.query.all():
        if drive.drive_type in buckets:
            buckets[drive.drive_type].add(drive.size_tb)
    return {bucket: sorted(sizes) for bucket, sizes in buckets.items()}


def _validated_ram_sizes():
    """RAM options for the validated picker: every size the hardware catalog
    offers (across all models) unioned with the standard baseline, so an
    admin-added RAM size shows up while the generic list never shrinks."""
    catalog = {
        row.size_gb
        for row in RamOption.query.with_entities(RamOption.size_gb).distinct()
    }
    return sorted(set(RAM_SIZES_GB) | catalog)


# The supported-language list and endonyms live in i18n.py (single source of truth,
# shared with the export translator) and are imported above.


def pick_lang():
    """Resolve the active UI language for the current request — READ ONLY, never
    writes a cookie. Precedence: the `lang` cookie (set only when the user
    explicitly picks a language) -> the browser's Accept-Language -> English. Used
    both to render pages and to language-match generated exports."""
    cookie = (request.cookies.get("lang") or "").lower()
    if cookie in SUPPORTED_LANGS:
        return cookie
    return request.accept_languages.best_match(SUPPORTED_LANGS) or "en"


def create_app():
    app = Flask(__name__)

    # Behind nginx: trust one proxy hop for the client IP (X-Forwarded-For) and
    # scheme (X-Forwarded-Proto) so rate limiting buckets per real client and
    # Flask knows requests are HTTPS. X-Forwarded-Host is deliberately NOT
    # trusted — email links use a configured base URL (see auth.app_base_url) to
    # avoid Host-header poisoning.
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=0)

    # Cap request bodies (defense against memory-exhaustion uploads). Generous
    # enough for a large RVTools/Live Optics export; the saved-config payload has
    # its own tighter 4 MB check in the configs blueprint.
    app.config["MAX_CONTENT_LENGTH"] = int(
        os.environ.get("MAX_CONTENT_LENGTH_BYTES", str(32 * 1024 * 1024)))

    # Signed-cookie sessions. SECRET_KEY must be set in production (and shared
    # across gunicorn workers, since each validates the same cookie signature).
    # Fall back to an ephemeral key for dev with a loud warning — sessions then
    # won't survive a restart or span multiple workers.
    secret = os.environ.get("SECRET_KEY")
    if not secret:
        # In production (signalled by SESSION_COOKIE_SECURE=true behind TLS) an
        # ephemeral key is dangerous: it churns logins across workers/restarts and
        # makes the Fernet-encrypted SMTP secret undecryptable. Refuse to boot.
        if os.environ.get("SESSION_COOKIE_SECURE", "false").lower() == "true":
            raise RuntimeError(
                "SECRET_KEY must be set in production. Refusing to start with an "
                "ephemeral key (SESSION_COOKIE_SECURE=true indicates prod)."
            )
        secret = secrets.token_hex(32)
        app.logger.warning(
            "SECRET_KEY not set — generated an ephemeral key. Logins will not "
            "persist across restarts or gunicorn workers. Set SECRET_KEY in prod."
        )
    app.config["SECRET_KEY"] = secret
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
    # Only require HTTPS for the cookie when explicitly told we're behind TLS.
    app.config["SESSION_COOKIE_SECURE"] = (
        os.environ.get("SESSION_COOKIE_SECURE", "false").lower() == "true"
    )

    init_db(app)
    limiter.init_app(app)

    # CSRF defence-in-depth: reject state-changing requests whose Origin/Referer
    # is a different host. SameSite=Lax already blocks cross-site cookie sending,
    # so this is a second layer. Fail-open when no Origin/Referer is present (non-
    # browser clients like curl) — those can't be a browser CSRF vector anyway.
    def _allowed_hosts():
        hosts = set()
        base = (os.environ.get("APP_BASE_URL") or "").strip()
        if base:
            h = urlparse(base if "://" in base else "https://" + base).hostname
            if h:
                hosts.add(h.lower())
        try:
            h = urlparse("http://" + request.host).hostname
            if h:
                hosts.add(h.lower())
        except Exception:
            pass
        return hosts

    @app.before_request
    def _csrf_origin_guard():
        if request.method in ("GET", "HEAD", "OPTIONS", "TRACE"):
            return
        src = request.headers.get("Origin") or request.headers.get("Referer")
        if not src:
            return
        try:
            host = urlparse(src).hostname
        except Exception:
            return
        if host and host.lower() not in _allowed_hosts():
            return jsonify({"error": "Cross-origin request blocked"}), 403

    register_auth(app)
    register_projects(app)
    app.register_blueprint(admin_bp)

    # Daily retention/GDPR-anonymization scheduler. Disabled (ENABLE_SCHEDULER=0)
    # for one-off processes like seeding/CLI; on by default for the web server.
    if os.environ.get("ENABLE_SCHEDULER", "1") != "0":
        start_scheduler(app)
        # Drains queued bundle exports. Claims are atomic, so the one thread
        # per gunicorn worker cooperate rather than build the same job several
        # times over (docs/projects-plan.md §7.2).
        from export_worker import start_export_worker
        start_export_worker(app)

    # Cache-bust static assets by file mtime so a rebuild always serves fresh
    # JS/CSS (no more stale-cache surprises during iteration).
    @app.context_processor
    def _asset_helper():
        def asset(path):
            full = os.path.join(app.static_folder, path)
            try:
                v = int(os.path.getmtime(full))
            except OSError:
                v = 0
            return f"/static/{path}?v={v}"
        return {"asset": asset}

    # UI language selection. Order of precedence:
    #   1. the `lang` cookie, but only if it names a supported language — this is
    #      written client-side (i18n.js setLang) ONLY when the user explicitly
    #      picks a language, so it always reflects a deliberate choice;
    #   2. otherwise the browser's Accept-Language header (auto-detection) —
    #      read-only, never persisted;
    #   3. English as the final fallback.
    # Every template gets `lang` (the active code) and `supported_langs` (for the
    # switcher); client-side i18n.js reads the code back from <html lang="..">.
    @app.context_processor
    def _lang_helper():
        return {"lang": pick_lang(), "supported_langs": SUPPORTED_LANGS,
                "lang_names": LANG_NAMES}

    @app.route("/")
    def index():
        # Surface the admin-tuned sizing defaults the client needs at load time
        # (the ratio slider's starting position). Resilient: if the settings read
        # fails, T keeps its last/default values so the page still renders.
        try:
            refresh_from_db()
        except Exception:
            pass
        return render_template("index.html", default_vcpu_ratio=T.default_vcpu_ratio,
                               max_day_one_storage_pct=T.max_day_one_storage_pct,
                               max_day_one_ram_pct=T.max_day_one_ram_pct)

    @app.route("/favicon.ico")
    def favicon():
        """Browsers ask for /favicon.ico unprompted, regardless of the <link>
        tags, and a 404 on every page load is noise in the console and the
        access log."""
        return send_file(
            os.path.join(app.static_folder, "img", "favicon.ico"),
            mimetype="image/vnd.microsoft.icon")

    @app.route("/privacy")
    def privacy():
        return render_template("privacy.html")

    @app.route("/api/models")
    def get_models():
        mode = request.args.get("mode", "appliance")
        status_filter = request.args.get("status", "active")

        if mode == "appliance":
            # The "Size For Model" picker must list exactly the models the
            # recommendation engine will consider for the chosen sizing mode
            # (see recommend.generate_recommendations): validated mode includes
            # validated-only platforms but drops NVMe+SSD (1+1) models, while
            # certified mode is the inverse.
            validated = request.args.get("sizing") == "validated"
            query = Model.query.options(
                joinedload(Model.cpu_links).joinedload(ModelCpuOption.cpu),
                joinedload(Model.nic_links).joinedload(ModelNicOption.nic),
                joinedload(Model.ram_options),
                joinedload(Model.storage_config)
                    .joinedload(StorageConfig.drive_links)
                    .joinedload(StorageConfigDrive.drive),
            )
            if status_filter == "active":
                query = query.filter(Model.status == "Active")
            elif status_filter == "all_current":
                query = query.filter(Model.status.in_(["Active", "EOL"]))
            if not validated:
                # Validated-only platforms have no certified equivalent, so they
                # don't belong in the certified appliance picker.
                query = query.filter(Model.validated_only == False)  # noqa: E712

            models = {}
            for m in query.order_by(Model.category, Model.name).all():
                # Validated mode can't use NVMe+SSD models (inherently 2-disk),
                # so exclude them to match what the engine will actually size.
                if validated and m.storage_config \
                        and m.storage_config.storage_type == "nvme_and_ssd":
                    continue
                models[m.name] = m.to_dict()
            return jsonify(models)
        else:
            nics = [n.to_dict() for n in ValidatedNic.query.all()]
            platforms = [p.to_dict() for p in ValidatedPlatform.query.filter_by(status="Active").all()]
            return jsonify({
                "nics": nics,
                "disk_sizes": _validated_disk_sizes(),
                "ram_sizes": _validated_ram_sizes(),
                "platforms": platforms,
            })

    @app.route("/api/calculate", methods=["POST"])
    def calculate():
        # Load the current admin-tuned overheads/limits for this request.
        refresh_from_db()
        data = request.json
        mode = data.get("mode", "appliance")
        # Coerce + bound node_count: it drives per-node loops and SVG generation,
        # so a non-int or absurd value must be rejected before any work.
        try:
            node_count = int(data.get("node_count", 3))
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid node count"}), 400

        if node_count < 1:
            return jsonify({"error": "Minimum 1 node required"}), 400
        if node_count > MAX_NODE_COUNT:
            return jsonify({"error": f"Node count must be {MAX_NODE_COUNT} or fewer"}), 400

        if mode == "appliance":
            return jsonify(calculate_appliance(data, node_count))
        else:
            return jsonify(calculate_validated(data, node_count))

    @app.route("/api/import-liveoptics", methods=["POST"])
    def import_liveoptics():
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        f = request.files["file"]
        if not f.filename or not f.filename.endswith(".xlsx"):
            return jsonify({"error": "File must be an .xlsx Excel file"}), 400
        # Content sniff: an .xlsx is a ZIP (starts with "PK\x03\x04"). Reject
        # anything else up front rather than feeding arbitrary bytes to openpyxl.
        head = f.stream.read(4)
        f.stream.seek(0)
        if head != b"PK\x03\x04":
            return jsonify({"error": "File must be an .xlsx Excel file"}), 400

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        try:
            f.save(tmp.name)
            tmp.close()

            file_digest = _file_sha256(tmp.name)
            file_type = _detect_file_type(tmp.name)
            if file_type == "rvtools":
                data = parse_rvtools(tmp.name)
            elif file_type == "liveoptics":
                data = parse_liveoptics(tmp.name)
            else:
                return jsonify({"error": "Unrecognised file format. Please upload a Live Optics or RVTools Excel export."}), 400

            vcpu_ratio = request.form.get("vcpu_ratio", type=float)
            result = generate_recommendations(data["summary"], vcpu_ratio)
            return jsonify({
                "summary": data["summary"],
                "project": data["project"],
                "hosts": data["hosts"],
                "datastores": data["datastores"],
                "vms": data["vms"],
                "vm_count": len(data["vms"]),
                "active_vm_count": data["summary"]["active_vms"],
                # Per-source-cluster summaries; empty unless the source held
                # more than one vSphere cluster. Drives the "size each cluster
                # separately" option in the UI.
                "clusters": data.get("clusters", []),
                "recommendations": result["recommendations"],
                "projection": result["projection"],
                "warnings": result.get("warnings", []),
                # Data-quality caveats about the import itself (no IOPS/perf,
                # assumed ratio, unclustered hosts, ...). {code, params} objects
                # translated client-side; shown on the wizard's Environment step.
                "import_warnings": build_import_warnings(data, file_type),
                "source": file_type,
                # Provenance for the project view and the export appendix
                # (docs/projects-plan.md §8). Combining several customer files
                # into one proposal is exactly when the reader needs to know
                # which number came from where. The digest also lets the client
                # warn when the same file is imported into a project twice.
                "source_meta": {
                    "file_name": f.filename,
                    "file_type": file_type,
                    "file_sha256": file_digest,
                    "imported_at": _utcnow_iso(),
                    "host_count": len(data.get("hosts") or []),
                    "vm_count": len(data.get("vms") or []),
                    "parser_version": PARSER_VERSION,
                },
            })
        except Exception as e:
            app.logger.warning("Import parse failed: %s", e)
            return jsonify({"error": "Could not parse the file. Upload a valid Live Optics or RVTools .xlsx export."}), 400
        finally:
            os.unlink(tmp.name)

    @app.route("/api/recommend", methods=["POST"])
    def recommend():
        data = request.json
        summary = data.get("summary")
        if not summary:
            return jsonify({"error": "No summary provided"}), 400
        vcpu_ratio = data.get("vcpu_ratio")
        growth_pct = data.get("growth_pct", 10)
        snapshot_pct = data.get("snapshot_pct", 20)
        years = data.get("years", 5)
        target_nodes = data.get("target_nodes")
        storage_pref = data.get("storage_pref")
        size_full_cluster = data.get("size_full_cluster", False)
        sizing_mode = data.get("sizing_mode", "certified")
        allow_storage_only = data.get("allow_storage_only", False)
        target_model = data.get("target_model")
        include_eol_eos = data.get("include_eol_eos", False)
        max_day_one_storage_pct = data.get("max_day_one_storage_pct")
        max_day_one_ram_pct = data.get("max_day_one_ram_pct")
        # Optional source-environment CPU benchmark for the perf comparison
        # (SPECrate2017 or PassMark, per the detected source CPU class).
        source_perf_index = data.get("source_perf_index")
        source_perf_type = data.get("source_perf_type")
        # Multi-site DR: inbound replication reserve this cluster must host, and
        # whether the compute reserve is held steady-state or only on failover.
        replication_reserve = data.get("replication_reserve")
        replication_compute_mode = data.get("replication_compute_mode", "reserved")
        allow_single_node = data.get("allow_single_node", False)
        result = generate_recommendations(summary, vcpu_ratio,
                                          growth_pct, snapshot_pct, years,
                                          target_nodes=target_nodes,
                                          storage_pref=storage_pref,
                                          size_full_cluster=size_full_cluster,
                                          sizing_mode=sizing_mode,
                                          allow_storage_only=allow_storage_only,
                                          target_model=target_model,
                                          include_eol_eos=include_eol_eos,
                                          max_day_one_storage_pct=max_day_one_storage_pct,
                                          max_day_one_ram_pct=max_day_one_ram_pct,
                                          source_perf_index=source_perf_index,
                                          source_perf_type=source_perf_type,
                                          replication_reserve=replication_reserve,
                                          replication_compute_mode=replication_compute_mode,
                                          allow_single_node=allow_single_node)
        return jsonify(result)

    @app.route("/api/cpu-perf")
    def cpu_perf():
        """Look up a CPU's benchmark score by (fuzzy) description, to auto-fill
        the source-benchmark field on import/manual. Per-CPU/socket value — the
        caller scales by the source socket count. Tries our curated appliance
        catalog first (precise), then the broad SPECrate2017 lookup (~625 CPUs
        averaged from all published SPEC CPU 2017 int-rate results) so arbitrary
        SOURCE CPUs resolve too. found=false only when neither knows it."""
        import cpu_benchmarks
        from cpu_specs import CPU_SPECS, cpu_model_key, perf_index as _perf_index
        q = request.args.get("q", "")
        spec = CPU_SPECS.get(cpu_model_key(q) or "")
        if spec:
            ptype = "specrate" if spec.get("specrate_int") is not None else "passmark"
            return jsonify({
                "found": True,
                "model": spec["model"],
                "perf_type": ptype,
                "perf_index": _perf_index(spec),
                "specrate_int": spec.get("specrate_int"),
                "passmark_cpu_mark": spec.get("passmark_cpu_mark"),
                "passmark_single": spec.get("passmark_single"),
                "source": "catalog",
            })
        hit = cpu_benchmarks.lookup(q)
        if hit:
            return jsonify({
                "found": True,
                "model": hit["model"],
                "perf_type": "specrate",
                "perf_index": hit["specrate_int"],
                "specrate_int": hit["specrate_int"],
                "passmark_cpu_mark": None,
                "passmark_single": None,
                "source": "spec-cpu2017",
                "samples": hit["samples"],
            })
        return jsonify({"found": False})

    @app.route("/api/export-config", methods=["POST"])
    @limiter.limit("20 per minute")
    @export_gate
    def export_config():
        data = request.json
        if not data:
            return jsonify({"error": "No data provided"}), 400
        if not _can_export_editable():
            return jsonify({"error": "The editable PowerPoint is available to Scale users only. Use the PDF instead."}), 403
        try:
            buf = generate_config_slide(data, lang=pick_lang())
            mode = data.get("mode", "config")
            model = data.get("model", mode)
            nodes = data.get("node_count", "")
            filename = f"SC_Config_{model}_{nodes}N.pptx"
            return send_file(buf, as_attachment=True, download_name=filename,
                             mimetype="application/vnd.openxmlformats-officedocument.presentationml.presentation")
        except Exception as e:
            app.logger.exception("Config slide generation failed: %s", e)
            return jsonify({"error": "Failed to generate the configuration slide."}), 500

    @app.route("/api/export-config-pdf", methods=["POST"])
    @limiter.limit("20 per minute")
    @export_gate
    def export_config_pdf():
        data = request.json
        if not data:
            return jsonify({"error": "No data provided"}), 400
        try:
            pptx_buf = generate_config_slide(data, lang=pick_lang())
            pdf = convert_pptx_to_pdf(pptx_buf.getvalue())
            if not pdf:
                return jsonify({"error": "PDF conversion is unavailable on this server."}), 503
            model = data.get("model", data.get("mode", "config"))
            nodes = data.get("node_count", "")
            filename = f"SC_Config_{model}_{nodes}N.pdf"
            return send_file(io.BytesIO(pdf), as_attachment=True, download_name=filename,
                             mimetype="application/pdf")
        except Exception as e:
            app.logger.exception("Config PDF generation failed: %s", e)
            return jsonify({"error": "Failed to generate the configuration PDF."}), 500

    @app.route("/api/export-proposal", methods=["POST"])
    @limiter.limit("20 per minute")
    @export_gate
    def export_proposal():
        data = request.json
        summary = data.get("summary")
        recommendation = data.get("recommendation")
        projection = data.get("projection")
        source_perf = data.get("source_perf")
        if not summary or not recommendation or not projection:
            return jsonify({"error": "Missing summary, recommendation, or projection"}), 400
        if not _can_export_editable():
            return jsonify({"error": "The editable PowerPoint is available to Scale users only. Use the PDF instead."}), 403

        try:
            buf = generate_proposal(summary, recommendation, projection, source_perf,
                                    lang=pick_lang())
            model_name = recommendation.get("model", "proposal")
            filename = f"SC_Proposal_{model_name}_{recommendation.get('node_count', '')}N.pptx"
            return send_file(buf, as_attachment=True, download_name=filename,
                             mimetype="application/vnd.openxmlformats-officedocument.presentationml.presentation")
        except Exception as e:
            app.logger.exception("Proposal generation failed: %s", e)
            return jsonify({"error": "Failed to generate the proposal."}), 500

    def _proposal_payload():
        data = request.json or {}
        return (data.get("summary"), data.get("recommendation"),
                data.get("projection"), data.get("source_perf"))

    def _can_export_editable():
        # Editable source files (Word, PPTX) are limited to Scale users and super
        # admins; everyone else is restricted to read-only PDFs.
        u = current_user()
        return bool(u and (u.is_scale or u.is_super_admin))

    @app.route("/api/export-docx", methods=["POST"])
    @limiter.limit("20 per minute")
    @export_gate
    def export_docx_route():
        summary, recommendation, projection, source_perf = _proposal_payload()
        if not summary or not recommendation or not projection:
            return jsonify({"error": "Missing summary, recommendation, or projection"}), 400
        if not _can_export_editable():
            return jsonify({"error": "The editable Word document is available to Scale users only. Use the PDF instead."}), 403
        try:
            buf = build_proposal_docx(summary, recommendation, projection, source_perf,
                                      lang=pick_lang())
            fn = f"SC_Proposal_{recommendation.get('model', 'proposal')}_{recommendation.get('node_count', '')}N.docx"
            return send_file(buf, as_attachment=True, download_name=fn,
                             mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        except Exception as e:
            app.logger.exception("Document generation failed: %s", e)
            return jsonify({"error": "Failed to generate the document."}), 500

    @app.route("/api/export-pdf", methods=["POST"])
    @limiter.limit("20 per minute")
    @export_gate
    def export_pdf_route():
        summary, recommendation, projection, source_perf = _proposal_payload()
        if not summary or not recommendation or not projection:
            return jsonify({"error": "Missing summary, recommendation, or projection"}), 400
        try:
            docx_buf = build_proposal_docx(summary, recommendation, projection, source_perf,
                                           lang=pick_lang())
            pdf = convert_docx_to_pdf(docx_buf.getvalue())
            if not pdf:
                return jsonify({"error": "PDF conversion is unavailable on this server."}), 503
            fn = f"SC_Proposal_{recommendation.get('model', 'proposal')}_{recommendation.get('node_count', '')}N.pdf"
            return send_file(io.BytesIO(pdf), as_attachment=True, download_name=fn,
                             mimetype="application/pdf")
        except Exception as e:
            app.logger.exception("PDF generation failed: %s", e)
            return jsonify({"error": "Failed to generate the PDF."}), 500

    @app.route("/api/export-presentation-pdf", methods=["POST"])
    @limiter.limit("20 per minute")
    @export_gate
    def export_presentation_pdf_route():
        summary, recommendation, projection, source_perf = _proposal_payload()
        if not summary or not recommendation or not projection:
            return jsonify({"error": "Missing summary, recommendation, or projection"}), 400
        try:
            pptx_buf = generate_proposal(summary, recommendation, projection, source_perf,
                                         lang=pick_lang())
            pdf = convert_pptx_to_pdf(pptx_buf.getvalue())
            if not pdf:
                return jsonify({"error": "PDF conversion is unavailable on this server."}), 503
            fn = f"SC_Presentation_{recommendation.get('model', 'proposal')}_{recommendation.get('node_count', '')}N.pdf"
            return send_file(io.BytesIO(pdf), as_attachment=True, download_name=fn,
                             mimetype="application/pdf")
        except Exception as e:
            app.logger.exception("Presentation PDF generation failed: %s", e)
            return jsonify({"error": "Failed to generate the presentation PDF."}), 500

    # ---- Combined multi-site exports (one document, per-cluster sections) ----
    def _bundle_payload():
        data = request.get_json(silent=True)
        if not isinstance(data, dict):
            return None
        clusters = data.get("clusters")
        if not isinstance(clusters, list) or not clusters:
            return None
        if len(clusters) > MAX_EXPORT_SECTIONS:
            return None
        for cl in clusters:
            if not (cl.get("summary") and cl.get("recommendation") and cl.get("projection")):
                return None
        return clusters

    @app.route("/api/export-bundle-proposal", methods=["POST"])
    @limiter.limit("20 per minute")
    @export_gate
    def export_bundle_proposal():
        clusters = _bundle_payload()
        if not clusters:
            return jsonify({"error": "Missing or incomplete cluster data"}), 400
        if not _can_export_editable():
            return jsonify({"error": "The editable PowerPoint is available to Scale users only. Use the PDF instead."}), 403
        try:
            buf = generate_bundle_proposal(clusters, lang=pick_lang())
            fn = f"SC_Proposal_MultiSite_{len(clusters)}clusters.pptx"
            return send_file(buf, as_attachment=True, download_name=fn,
                             mimetype="application/vnd.openxmlformats-officedocument.presentationml.presentation")
        except Exception as e:
            app.logger.exception("Multi-site proposal generation failed: %s", e)
            return jsonify({"error": "Failed to generate the proposal."}), 500

    @app.route("/api/export-bundle-docx", methods=["POST"])
    @limiter.limit("20 per minute")
    @export_gate
    def export_bundle_docx():
        clusters = _bundle_payload()
        if not clusters:
            return jsonify({"error": "Missing or incomplete cluster data"}), 400
        if not _can_export_editable():
            return jsonify({"error": "The editable Word document is available to Scale users only. Use the PDF instead."}), 403
        try:
            buf = build_bundle_proposal_docx(clusters, lang=pick_lang())
            fn = f"SC_Proposal_MultiSite_{len(clusters)}clusters.docx"
            return send_file(buf, as_attachment=True, download_name=fn,
                             mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        except Exception as e:
            app.logger.exception("Multi-site document generation failed: %s", e)
            return jsonify({"error": "Failed to generate the document."}), 500

    @app.route("/api/export-bundle-pdf", methods=["POST"])
    @limiter.limit("20 per minute")
    @export_gate
    def export_bundle_pdf():
        clusters = _bundle_payload()
        if not clusters:
            return jsonify({"error": "Missing or incomplete cluster data"}), 400
        try:
            docx_buf = build_bundle_proposal_docx(clusters, lang=pick_lang())
            pdf = convert_docx_to_pdf(docx_buf.getvalue())
            if not pdf:
                return jsonify({"error": "PDF conversion is unavailable on this server."}), 503
            fn = f"SC_Proposal_MultiSite_{len(clusters)}clusters.pdf"
            return send_file(io.BytesIO(pdf), as_attachment=True, download_name=fn,
                             mimetype="application/pdf")
        except Exception as e:
            app.logger.exception("Multi-site PDF generation failed: %s", e)
            return jsonify({"error": "Failed to generate the PDF."}), 500

    @app.route("/api/export-bundle-presentation-pdf", methods=["POST"])
    @limiter.limit("20 per minute")
    @export_gate
    def export_bundle_presentation_pdf():
        clusters = _bundle_payload()
        if not clusters:
            return jsonify({"error": "Missing or incomplete cluster data"}), 400
        try:
            pptx_buf = generate_bundle_proposal(clusters, lang=pick_lang())
            pdf = convert_pptx_to_pdf(pptx_buf.getvalue())
            if not pdf:
                return jsonify({"error": "PDF conversion is unavailable on this server."}), 503
            fn = f"SC_Presentation_MultiSite_{len(clusters)}clusters.pdf"
            return send_file(io.BytesIO(pdf), as_attachment=True, download_name=fn,
                             mimetype="application/pdf")
        except Exception as e:
            app.logger.exception("Multi-site presentation PDF generation failed: %s", e)
            return jsonify({"error": "Failed to generate the presentation PDF."}), 500

    return app


def _file_sha256(file_path):
    """Digest of the uploaded file, kept as provenance. Only the hash is stored
    — never the file itself — so a re-import can be spotted without holding
    customer data (§8)."""
    import hashlib
    digest = hashlib.sha256()
    with open(file_path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _utcnow_iso():
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).isoformat()


def _detect_file_type(file_path):
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True)
    sheets = set(wb.sheetnames)
    wb.close()
    if "vInfo" in sheets or "vMetaData" in sheets:
        return "rvtools"
    if "ESX Hosts" in sheets or "Details" in sheets:
        return "liveoptics"
    return None


app = create_app()

if __name__ == "__main__":
    # Debug (Werkzeug interactive debugger = RCE) is OFF unless explicitly enabled
    # via FLASK_DEBUG=1 for local dev. Production runs under gunicorn, not this.
    app.run(host=os.environ.get("FLASK_RUN_HOST", "0.0.0.0"), port=5000,
            debug=os.environ.get("FLASK_DEBUG", "") == "1")
