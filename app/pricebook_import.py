"""Feed B — the licence-pricing import (docs/pricebook-plan.md §5.2).

Narrow and defensible: `HCOS-*` rows only. Six columns, one product family, no
hardware pricing at all. The hardware side of the price list is deliberately not
read — see §9 for why.

Parser rules, all of which exist because a real export moved on us once:

  * Build the column index **by header name**, never by position.
  * Classify by SKU prefix; never infer meaning from the SKU family digit, which
    is a carrier/generation code with no published mapping.
  * **Unmatched rows are counted and reported, never silently dropped.** That
    count is the early warning that the export format changed.

Validated against `_archive/Scale Computing Q4 2025 EUR Master Price List.xlsx`:
370 licence rows -> 360 banded, 9 flat, 1 unmatched. The single unmatched row
is `HCOS-S-POC`, the €0 60-day proof-of-concept licence — correctly excluded,
since it is neither a band nor a priced flat SKU.
"""
import re

from xlsx_utils import sheet_rows as _sheet_rows

import licensing

LICENSE_FAMILY = "HyperCore License and Support"

COL_CODE = "Product: Product Code"
COL_DESC = "Product: Product Description"
COL_PRICE = "List Price"
COL_CURRENCY = "Currency"
COL_FAMILY = "Product: Product Family"

REQUIRED_COLUMNS = (COL_CODE, COL_DESC, COL_PRICE, COL_CURRENCY, COL_FAMILY)

# HCOS-<edition>-<term>-<cores>C[-PS|-SS].  [A-Z]+ on the edition on purpose: a
# future edition letter lands as data marked "priced, not yet selectable"
# rather than as an unmatched row.
BANDED_RE = re.compile(r"^HCOS-([A-Z]+)-(\d)-(\d+)C(?:-(PS|SS))?$")

# HCOS-<term>-{PE,SE}
FLAT_ESSENTIALS_RE = re.compile(r"^HCOS-(\d)-(PE|SE)$")

# HCOS-<term>-1S-<n>WL — single-site workload-count licence. Stored, not offered.
FLAT_SITE_RE = re.compile(r"^HCOS-(\d)-1S-(\d+)WL$")


class PricebookFormatError(ValueError):
    """The sheet is not shaped like a price list. Blocks apply (§4.4)."""


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_license_rows(rows):
    """Parse pre-read sheet rows into bands, flats and an unmatched report.

    Returns a dict:
        bands      [{edition, term_years, support_tier, core_band, price, sku}]
        flats      [{sku, kind, term_years, price, max_nodes, max_ram_gb,
                     workloads}]
        unmatched  [{sku, description}]   — surfaced, never dropped
        currency   the single currency seen, or None if mixed/absent
        counts     {"license_rows": n, "banded": n, "flat": n, "unmatched": n}
    """
    bands, flats, unmatched = [], [], []
    currencies = set()

    license_rows = [r for r in rows
                    if str(r.get(COL_FAMILY) or "").strip() == LICENSE_FAMILY]

    for r in license_rows:
        sku = str(r.get(COL_CODE) or "").strip()
        desc = str(r.get(COL_DESC) or "").strip()
        price = _num(r.get(COL_PRICE))
        cur = str(r.get(COL_CURRENCY) or "").strip().upper()
        if cur:
            currencies.add(cur)

        m = BANDED_RE.match(sku)
        if m:
            edition, term, cores, support = m.groups()
            bands.append({
                "sku": sku,
                "edition": edition,
                "term_years": int(term),
                # Editions with a single ladder omit the suffix; they are
                # Standard Support.
                "support_tier": support or licensing.SUPPORT_STANDARD,
                "core_band": int(cores),
                "price": price,
            })
            continue

        m = FLAT_ESSENTIALS_RE.match(sku)
        if m:
            term, kind = m.groups()
            flats.append({
                "sku": sku,
                "kind": kind,
                "term_years": int(term),
                "price": price,
                # Ceilings are product policy, not carried in the feed.
                "max_nodes": licensing.ESSENTIALS_EXACT_NODES,
                "max_ram_gb": licensing.ESSENTIALS_MAX_RAM_GB_PER_NODE,
                "workloads": None,
            })
            continue

        m = FLAT_SITE_RE.match(sku)
        if m:
            term, workloads = m.groups()
            flats.append({
                "sku": sku,
                "kind": licensing.FLAT_SITE_WORKLOAD,
                "term_years": int(term),
                "price": price,
                "max_nodes": None,
                "max_ram_gb": None,
                "workloads": int(workloads),
            })
            continue

        # Everything else in the family: PoC SKUs, anything new. Recorded so the
        # count is visible in the diff report.
        unmatched.append({"sku": sku, "description": desc})

    return {
        "bands": bands,
        "flats": flats,
        "unmatched": unmatched,
        "currency": currencies.pop() if len(currencies) == 1 else None,
        "counts": {
            "license_rows": len(license_rows),
            "banded": len(bands),
            "flat": len(flats),
            "unmatched": len(unmatched),
        },
    }


def parse_pricebook(file_path, sheet=None):
    """Read an xlsx price list and parse its licence rows.

    Sanity gates that block apply (§4.4): required headers missing, or zero
    licence rows parsed. Both mean the export format moved and the safe action
    is to fail loudly rather than apply a half-read feed.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        name = sheet or wb.sheetnames[0]
        rows = _sheet_rows(wb, name)
    finally:
        wb.close()

    if rows:
        missing = [c for c in REQUIRED_COLUMNS if c not in rows[0]]
        if missing:
            raise PricebookFormatError(
                "price list is missing required column(s): "
                + ", ".join(missing)
            )

    result = parse_license_rows(rows)
    if result["counts"]["license_rows"] == 0:
        raise PricebookFormatError(
            f"no rows in the '{LICENSE_FAMILY}' product family — this does not "
            "look like a Scale price list, or the family name changed"
        )
    return result


def build_book(parsed, region=None, feed_label=None):
    """Turn a parse result into a `licensing.LicenseBook`.

    Used by the seeder and by tests, so the pricing math can be exercised
    straight from a file without a database round-trip.
    """
    bands = {}
    for b in parsed["bands"]:
        if b["price"] is None:
            continue
        key = (b["edition"], b["term_years"], b["support_tier"])
        bands.setdefault(key, {})[b["core_band"]] = b["price"]

    flats = {}
    for f in parsed["flats"]:
        if f["price"] is None:
            continue
        flats[(f["kind"], f["term_years"])] = f["price"]

    return licensing.LicenseBook(
        bands=bands, flats=flats, region=region, feed_label=feed_label,
        currency=parsed.get("currency") or "EUR",
    )


# ── Seeding ──────────────────────────────────────────────────────────────────
# Slice 1 seeds the tables by running this parser over the archived price list.
# The admin upload -> diff -> apply flow (§4.4) is deferred: feed B does not
# exist from Salesforce yet, so building an importer for a guessed file shape
# would be work we would throw away. The tables and the lookup are the eventual
# home either way, so nothing here is temporary.

DEFAULT_REGION = "EMEA"

# Seeded eligibility, in the closed vocabulary. These are product policy, not
# feed data — the feed carries prices, these carry shape. Editing a value is an
# admin row edit; a new KIND of constraint is a code change (§11).
SEEDED_RULES = {
    licensing.EDITION_STANDARD: {
        "selectable": True,
        "rule": {"bundleable": True, "role_gated": False},
    },
    licensing.EDITION_BRS: {
        # Deferred (§6). Priced and stored, never offered.
        "selectable": False,
        "rule": {"requires_single_node": True, "role_gated": True,
                 "bundleable": False},
    },
    licensing.EDITION_VIDEO: {
        "selectable": False,
        "rule": {"min_hci_nodes": 2, "role_gated": True,
                 "workload_class": "cctv", "bundleable": False},
    },
    licensing.FLAT_ESSENTIALS: {
        "selectable": True,
        "rule": {"exact_nodes": licensing.ESSENTIALS_EXACT_NODES,
                 "max_ram_gb_per_node": licensing.ESSENTIALS_MAX_RAM_GB_PER_NODE,
                 "bundleable": False, "role_gated": False},
    },
    licensing.FLAT_PRO_ESSENTIALS: {
        "selectable": True,
        "rule": {"exact_nodes": licensing.ESSENTIALS_EXACT_NODES,
                 "max_ram_gb_per_node": licensing.ESSENTIALS_MAX_RAM_GB_PER_NODE,
                 "bundleable": False, "role_gated": False},
    },
}


def _sha256(path):
    import hashlib
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def seed_feed_from_file(file_path, region=DEFAULT_REGION, label=None,
                        effective_date=None, uploaded_by=None, sheet=None):
    """Parse a price list and store it as the current feed for `region`.

    Idempotent by content: re-seeding the same file for the same region is a
    no-op, so this is safe to call on every boot. A different file supersedes
    the previous feed rather than replacing it — older feeds are retained so a
    saved sizing stamped with a feed id still reproduces.

    Returns (feed, parsed) or (existing_feed, None) when already current.
    """
    import os
    from datetime import datetime, timezone

    from database import db
    from orm_models import (CatalogFeed, PriceLicenseBand, PriceLicenseFlat,
                            PriceLicenseRule)

    digest = _sha256(file_path)
    existing = CatalogFeed.query.filter_by(region=region,
                                           source_sha256=digest).first()
    if existing is not None:
        if not existing.is_current:
            CatalogFeed.query.filter_by(region=region, is_current=True) \
                .update({"is_current": False})
            existing.is_current = True
            db.session.commit()
        return existing, None

    parsed = parse_pricebook(file_path, sheet=sheet)

    CatalogFeed.query.filter_by(region=region, is_current=True) \
        .update({"is_current": False})

    feed = CatalogFeed(
        region=region,
        label=label or os.path.basename(file_path),
        currency=parsed.get("currency") or "EUR",
        source_sha256=digest,
        source_filename=os.path.basename(file_path),
        effective_date=effective_date,
        uploaded_by=uploaded_by,
        uploaded_at=datetime.now(timezone.utc),
        is_current=True,
        unmatched_count=parsed["counts"]["unmatched"],
    )
    db.session.add(feed)
    db.session.flush()

    for b in parsed["bands"]:
        if b["price"] is None:
            continue
        db.session.add(PriceLicenseBand(
            catalog_feed_id=feed.id, sku=b["sku"], edition=b["edition"],
            term_years=b["term_years"], support_tier=b["support_tier"],
            core_band=b["core_band"], price=b["price"]))

    for f in parsed["flats"]:
        if f["price"] is None:
            continue
        db.session.add(PriceLicenseFlat(
            catalog_feed_id=feed.id, sku=f["sku"], kind=f["kind"],
            term_years=f["term_years"], price=f["price"],
            max_nodes=f["max_nodes"], max_ram_gb=f["max_ram_gb"],
            workloads=f["workloads"]))

    for edition, spec in SEEDED_RULES.items():
        rule = licensing.validate_rule(dict(spec["rule"]))
        db.session.add(PriceLicenseRule(
            catalog_feed_id=feed.id, edition=edition,
            selectable=spec["selectable"], **rule))

    db.session.commit()
    return feed, parsed
