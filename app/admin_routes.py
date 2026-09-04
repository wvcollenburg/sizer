"""Admin blueprint – model CRUD + Excel import/export."""
import io
import tempfile
import os

from flask import Blueprint, render_template, jsonify, request, send_file, redirect, current_app
from sqlalchemy.orm import joinedload
from database import db
from auth import current_user
from orm_models import (
    Model, RamOption, StorageConfig,
    CpuCatalog, NicCatalog, DriveCatalog, DriveTypeIops, SizingSetting,
    ModelCpuOption, ModelNicOption, StorageConfigDrive,
)
from tunables import TUNABLE_DEFS, DEFAULTS as TUNABLE_DEFAULTS
from xlsx_utils import sheet_rows as _sheet_rows, parse_quantity as _parse_quantity

DRIVE_IOPS_TYPES = ["HDD", "SSD", "NVMe"]
_TUNABLE_BY_KEY = {d["key"]: d for d in TUNABLE_DEFS}

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


@admin_bp.before_request
def require_super_admin():
    """The entire models/catalog area is super-admin-only. Browser visits to a
    page are redirected to the main app (which hosts the login modal); API calls
    get a JSON 403."""
    user = current_user()
    if user is not None and user.is_super_admin:
        return None
    wants_json = (
        request.path.startswith("/admin/api/")
        or "application/json" in (request.headers.get("Accept") or "")
    )
    if wants_json:
        return jsonify({"error": "Super admin access required"}), 403
    return redirect("/?admin=1")


def _model_query():
    return Model.query.options(
        joinedload(Model.cpu_links).joinedload(ModelCpuOption.cpu),
        joinedload(Model.nic_links).joinedload(ModelNicOption.nic),
        joinedload(Model.ram_options),
        joinedload(Model.storage_config)
            .joinedload(StorageConfig.drive_links)
            .joinedload(StorageConfigDrive.drive),
    )


@admin_bp.route("/")
def admin_page():
    return render_template("admin.html")


# ── Catalog endpoints ────────────────────────────────────────────────────────

@admin_bp.route("/api/cpus")
def list_cpus():
    cpus = CpuCatalog.query.order_by(CpuCatalog.cores, CpuCatalog.ghz).all()
    result = []
    for c in cpus:
        used = ModelCpuOption.query.filter_by(cpu_id=c.id).count()
        result.append({"id": c.id, **c.to_dict(), "used_by": used})
    return jsonify(result)


def _cpu_num(v):
    try:
        return float(v) if v is not None and v != "" else None
    except (TypeError, ValueError):
        return None


def _cpu_int(v):
    try:
        return int(v) if v is not None and v != "" else None
    except (TypeError, ValueError):
        return None


def _apply_cpu_specs(cpu, data):
    """Set the optional generation/clock/core/benchmark columns from a
    create or update payload. Only touches keys actually present, so a partial
    update never wipes columns it didn't send."""
    if "generation" in data:
        cpu.generation = (data["generation"] or None)
    for key in ("p_cores", "e_cores", "passmark_cpu_mark"):
        if key in data:
            setattr(cpu, key, _cpu_int(data[key]))
    for key in ("base_ghz", "all_core_turbo_ghz", "max_turbo_ghz", "specrate_int"):
        if key in data:
            setattr(cpu, key, _cpu_num(data[key]))


@admin_bp.route("/api/cpus", methods=["POST"])
def create_cpu():
    data = request.json
    if not data or not data.get("desc"):
        return jsonify({"error": "Description is required"}), 400
    if CpuCatalog.query.filter_by(description=data["desc"]).first():
        return jsonify({"error": "CPU already exists"}), 409
    cpu = CpuCatalog(description=data["desc"], cores=int(data.get("cores", 0)),
                     threads=int(data.get("threads", 0)), ghz=float(data.get("ghz", 0)))
    _apply_cpu_specs(cpu, data)
    db.session.add(cpu)
    db.session.commit()
    return jsonify({"id": cpu.id, **cpu.to_dict()}), 201


@admin_bp.route("/api/cpus/<int:cpu_id>", methods=["PUT"])
def update_cpu(cpu_id):
    cpu = db.get_or_404(CpuCatalog, cpu_id)
    data = request.json
    if data.get("desc") and data["desc"] != cpu.description:
        if CpuCatalog.query.filter_by(description=data["desc"]).first():
            return jsonify({"error": "CPU already exists"}), 409
    cpu.description = data.get("desc", cpu.description)
    cpu.cores = int(data.get("cores", cpu.cores))
    cpu.threads = int(data.get("threads", cpu.threads))
    cpu.ghz = float(data.get("ghz", cpu.ghz))
    _apply_cpu_specs(cpu, data)
    db.session.commit()
    return jsonify({"id": cpu.id, **cpu.to_dict()})


@admin_bp.route("/api/cpus/<int:cpu_id>", methods=["DELETE"])
def delete_cpu(cpu_id):
    cpu = db.get_or_404(CpuCatalog, cpu_id)
    used = ModelCpuOption.query.filter_by(cpu_id=cpu.id).count()
    if used:
        return jsonify({"error": f"CPU is used by {used} model(s). Remove it from those models first."}), 409
    db.session.delete(cpu)
    db.session.commit()
    return jsonify({"message": "CPU deleted"})


@admin_bp.route("/api/nics")
def list_nics():
    nics = NicCatalog.query.order_by(NicCatalog.speed, NicCatalog.description).all()
    result = []
    for n in nics:
        used = ModelNicOption.query.filter_by(nic_id=n.id).count()
        result.append({"id": n.id, **n.to_dict(), "used_by": used})
    return jsonify(result)


@admin_bp.route("/api/nics", methods=["POST"])
def create_nic():
    data = request.json
    if not data or not data.get("desc"):
        return jsonify({"error": "Description is required"}), 400
    if NicCatalog.query.filter_by(description=data["desc"]).first():
        return jsonify({"error": "NIC already exists"}), 409
    nic = NicCatalog(description=data["desc"], ports=int(data.get("ports", 0)),
                     speed=data.get("speed", ""))
    db.session.add(nic)
    db.session.commit()
    return jsonify({"id": nic.id, **nic.to_dict()}), 201


@admin_bp.route("/api/nics/<int:nic_id>", methods=["PUT"])
def update_nic(nic_id):
    nic = db.get_or_404(NicCatalog, nic_id)
    data = request.json
    if data.get("desc") and data["desc"] != nic.description:
        if NicCatalog.query.filter_by(description=data["desc"]).first():
            return jsonify({"error": "NIC already exists"}), 409
    nic.description = data.get("desc", nic.description)
    nic.ports = int(data.get("ports", nic.ports))
    nic.speed = data.get("speed", nic.speed)
    db.session.commit()
    return jsonify({"id": nic.id, **nic.to_dict()})


@admin_bp.route("/api/nics/<int:nic_id>", methods=["DELETE"])
def delete_nic(nic_id):
    nic = db.get_or_404(NicCatalog, nic_id)
    used = ModelNicOption.query.filter_by(nic_id=nic.id).count()
    if used:
        return jsonify({"error": f"NIC is used by {used} model(s). Remove it from those models first."}), 409
    db.session.delete(nic)
    db.session.commit()
    return jsonify({"message": "NIC deleted"})


@admin_bp.route("/api/drives")
def list_drives():
    drives = DriveCatalog.query.order_by(DriveCatalog.drive_type, DriveCatalog.size_tb).all()
    result = []
    for d in drives:
        used = StorageConfigDrive.query.filter_by(drive_id=d.id).count()
        result.append({"id": d.id, "drive_type": d.drive_type, "size_tb": d.size_tb, "used_by": used})
    return jsonify(result)


@admin_bp.route("/api/drives", methods=["POST"])
def create_drive():
    data = request.json
    dtype = data.get("drive_type", "")
    size = float(data.get("size_tb", 0))
    if not dtype or size <= 0:
        return jsonify({"error": "Drive type and size are required"}), 400
    if DriveCatalog.query.filter_by(drive_type=dtype, size_tb=size).first():
        return jsonify({"error": "Drive already exists"}), 409
    drive = DriveCatalog(drive_type=dtype, size_tb=size)
    db.session.add(drive)
    db.session.commit()
    return jsonify({"id": drive.id, "drive_type": drive.drive_type, "size_tb": drive.size_tb}), 201


@admin_bp.route("/api/drives/<int:drive_id>", methods=["PUT"])
def update_drive(drive_id):
    drive = db.get_or_404(DriveCatalog, drive_id)
    data = request.json
    new_type = data.get("drive_type", drive.drive_type)
    new_size = float(data.get("size_tb", drive.size_tb))
    if (new_type != drive.drive_type or new_size != drive.size_tb):
        if DriveCatalog.query.filter_by(drive_type=new_type, size_tb=new_size).first():
            return jsonify({"error": "Drive already exists"}), 409
    drive.drive_type = new_type
    drive.size_tb = new_size
    db.session.commit()
    return jsonify({"id": drive.id, "drive_type": drive.drive_type, "size_tb": drive.size_tb})


@admin_bp.route("/api/drives/<int:drive_id>", methods=["DELETE"])
def delete_drive(drive_id):
    drive = db.get_or_404(DriveCatalog, drive_id)
    used = StorageConfigDrive.query.filter_by(drive_id=drive.id).count()
    if used:
        return jsonify({"error": f"Drive is used by {used} storage config(s). Remove it from those models first."}), 409
    db.session.delete(drive)
    db.session.commit()
    return jsonify({"message": "Drive deleted"})


# ── Per-drive-type IOPS (configurable) ───────────────────────────────────────

@admin_bp.route("/api/drive-iops")
def list_drive_iops():
    rows = {r.drive_type: r for r in DriveTypeIops.query.all()}
    # Return in a stable, known order regardless of insertion order.
    return jsonify([rows[t].to_dict() for t in DRIVE_IOPS_TYPES if t in rows])


@admin_bp.route("/api/drive-iops", methods=["PUT"])
def update_drive_iops():
    data = request.json or {}
    # Accept either {"HDD": n, ...} or [{"drive_type": "HDD", "iops": n}, ...].
    if isinstance(data, list):
        data = {d.get("drive_type"): d.get("iops") for d in data}

    updates = {}
    for dtype in DRIVE_IOPS_TYPES:
        if dtype not in data:
            continue
        try:
            val = int(data[dtype])
        except (TypeError, ValueError):
            return jsonify({"error": f"{dtype} IOPS must be a whole number"}), 400
        if val < 0:
            return jsonify({"error": f"{dtype} IOPS cannot be negative"}), 400
        updates[dtype] = val

    if not updates:
        return jsonify({"error": "No valid IOPS values provided"}), 400

    for dtype, val in updates.items():
        row = DriveTypeIops.query.filter_by(drive_type=dtype).first()
        if row:
            row.iops = val
        else:
            db.session.add(DriveTypeIops(drive_type=dtype, iops=val))
    db.session.commit()
    return jsonify({"message": "Drive IOPS updated",
                    "drive_iops": [r.to_dict() for r in DriveTypeIops.query.all()]})


# ── Cluster-level IOPS sizing config (configurable) ──────────────────────────

def _sizing_config_dict():
    return {s.key: s.value for s in SizingSetting.query.all()}


@admin_bp.route("/api/sizing-config")
def get_sizing_config():
    return jsonify(_sizing_config_dict())


@admin_bp.route("/api/sizing-config", methods=["PUT"])
def update_sizing_config():
    data = request.json or {}
    # (key, parser/validator) — each returns the stored float or raises ValueError.
    def frac(v):
        f = float(v)
        if not 0 <= f <= 1:
            raise ValueError
        return f

    def derate(v):
        f = float(v)
        if not 0 <= f < 0.9:
            raise ValueError
        return f

    def rf(v):
        f = int(v)
        if f < 1:
            raise ValueError
        return float(f)

    validators = {
        "iops_derating_pct": (derate, "Derating must be between 0 and 0.9"),
        "iops_replication_factor": (rf, "Replication factor must be a whole number ≥ 1"),
        "iops_read_fraction": (frac, "Read fraction must be between 0 and 1"),
    }

    updates = {}
    for key, (parse, msg) in validators.items():
        if key not in data:
            continue
        try:
            updates[key] = parse(data[key])
        except (TypeError, ValueError):
            return jsonify({"error": msg}), 400

    if not updates:
        return jsonify({"error": "No valid sizing settings provided"}), 400

    for key, value in updates.items():
        row = SizingSetting.query.filter_by(key=key).first()
        if row:
            row.value = value
        else:
            db.session.add(SizingSetting(key=key, value=value))
    db.session.commit()
    return jsonify({"message": "Sizing config updated", "sizing_config": _sizing_config_dict()})


# ── Scoring / sizing / topology tunables ─────────────────────────────────────
# Reuses the SizingSetting key/value table; the field set + grouping is driven
# by tunables.TUNABLE_DEFS so the admin page renders itself from metadata.

def _tunable_values():
    """Current tunable values: defaults overlaid with any saved overrides,
    coerced to each tunable's declared type (ints stay ints)."""
    saved = {s.key: s.value for s in SizingSetting.query.all()}
    out = {}
    for d in TUNABLE_DEFS:
        key = d["key"]
        val = saved.get(key, d["default"])
        out[key] = int(round(val)) if d["type"] == "int" else float(val)
    return out


def _coerce_tunable(d, value):
    """Validate+coerce one incoming tunable value against its metadata.
    Returns the stored value or raises ValueError with a user message."""
    try:
        num = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{d['label']} must be a number")
    if d["type"] == "int":
        num = int(round(num))
    lo, hi = d.get("min"), d.get("max")
    if lo is not None and num < lo:
        raise ValueError(f"{d['label']} must be ≥ {lo}")
    if hi is not None and num > hi:
        raise ValueError(f"{d['label']} must be ≤ {hi}")
    return num


@admin_bp.route("/api/tunables")
def get_tunables():
    return jsonify({"defs": TUNABLE_DEFS, "values": _tunable_values()})


@admin_bp.route("/api/tunables", methods=["PUT"])
def update_tunables():
    data = request.json or {}
    updates = {}
    for key, raw in data.items():
        d = _TUNABLE_BY_KEY.get(key)
        if d is None:
            continue  # ignore unknown keys
        try:
            updates[key] = _coerce_tunable(d, raw)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400

    # Cross-field sanity: hybrid flash floor must not exceed its ceiling.
    lo = updates.get("hybrid_flash_min_pct", _tunable_values()["hybrid_flash_min_pct"])
    hi = updates.get("hybrid_flash_max_pct", _tunable_values()["hybrid_flash_max_pct"])
    if lo > hi:
        return jsonify({"error": "Hybrid flash floor must not exceed the ceiling"}), 400

    if not updates:
        return jsonify({"error": "No valid tunables provided"}), 400

    for key, value in updates.items():
        row = SizingSetting.query.filter_by(key=key).first()
        if row:
            row.value = value
        else:
            db.session.add(SizingSetting(key=key, value=value))
    db.session.commit()
    return jsonify({"message": "Tunables updated", "values": _tunable_values()})


@admin_bp.route("/api/tunables/reset", methods=["POST"])
def reset_tunables():
    for key, value in TUNABLE_DEFAULTS.items():
        row = SizingSetting.query.filter_by(key=key).first()
        if row:
            row.value = value
        else:
            db.session.add(SizingSetting(key=key, value=value))
    db.session.commit()
    return jsonify({"message": "Tunables reset to defaults", "values": _tunable_values()})


# ── List all models ──────────────────────────────────────────────────────────

@admin_bp.route("/api/models")
def list_models():
    models = _model_query().order_by(Model.category, Model.name).all()
    result = []
    for m in models:
        d = m.to_dict(include_internal=True)
        d["id"] = m.id
        d["name"] = m.name
        result.append(d)
    return jsonify(result)


# ── Get single model ─────────────────────────────────────────────────────────

@admin_bp.route("/api/models/<int:model_id>")
def get_model(model_id):
    m = _model_query().get_or_404(model_id)
    d = m.to_dict(include_internal=True)
    d["id"] = m.id
    d["name"] = m.name
    d["cpu_options"] = [
        {**link.cpu.to_dict(), "qty": link.quantity, "desc": link.cpu.description}
        for link in m.cpu_links
    ]
    d["nic_options"] = [
        {**link.nic.to_dict(), "qty": link.quantity, "desc": link.nic.description}
        for link in m.nic_links
    ]
    return jsonify(d)


# ── Create model ─────────────────────────────────────────────────────────────

@admin_bp.route("/api/models", methods=["POST"])
def create_model():
    data = request.json
    if not data or not data.get("name"):
        return jsonify({"error": "Model name is required"}), 400

    if Model.query.filter_by(name=data["name"]).first():
        return jsonify({"error": f"Model '{data['name']}' already exists"}), 409

    model = _build_model(data)
    db.session.add(model)
    db.session.commit()
    return jsonify({"id": model.id, "message": f"Model '{model.name}' created"}), 201


# ── Update model ─────────────────────────────────────────────────────────────

@admin_bp.route("/api/models/<int:model_id>", methods=["PUT"])
def update_model(model_id):
    model = db.get_or_404(Model, model_id)
    data = request.json

    if data.get("name") and data["name"] != model.name:
        if Model.query.filter_by(name=data["name"]).first():
            return jsonify({"error": f"Model '{data['name']}' already exists"}), 409

    model.name = data.get("name", model.name)
    model.status = data.get("status", model.status)
    model.category = data.get("category", model.category)
    model.form_factor = data.get("form_factor", model.form_factor)
    model.chassis = data.get("chassis", model.chassis)
    model.socket = data.get("socket", model.socket)
    model.psu = data.get("psu", model.psu)
    model.ram_slots = data.get("ram_slots", model.ram_slots)
    model.min_nodes = data.get("min_nodes", model.min_nodes)
    if "cost_tier" in data and data["cost_tier"] not in (None, ""):
        model.cost_tier = float(data["cost_tier"])
    if "validated_only" in data:
        model.validated_only = bool(data["validated_only"])
    model.notes = data.get("notes", model.notes)

    if "cpu_options" in data:
        ModelCpuOption.query.filter_by(model_id=model.id).delete()
        for i, cpu_data in enumerate(data["cpu_options"]):
            qty, base_desc = _parse_quantity(cpu_data["desc"])
            qty = cpu_data.get("qty", qty)
            cpu = _get_or_create_cpu(base_desc, cpu_data["cores"],
                                     cpu_data["threads"], cpu_data["ghz"])
            db.session.add(ModelCpuOption(
                model_id=model.id, cpu_id=cpu.id,
                quantity=qty, sort_order=i,
            ))

    if "ram_options_gb" in data:
        RamOption.query.filter_by(model_id=model.id).delete()
        for size in data["ram_options_gb"]:
            db.session.add(RamOption(model_id=model.id, size_gb=size))

    if "storage" in data:
        if model.storage_config:
            StorageConfigDrive.query.filter_by(
                storage_config_id=model.storage_config.id
            ).delete()
            db.session.delete(model.storage_config)
            db.session.flush()
        _add_storage(model, data["storage"])

    if "nic_options" in data:
        ModelNicOption.query.filter_by(model_id=model.id).delete()
        for i, nic_data in enumerate(data["nic_options"]):
            qty, base_desc = _parse_quantity(nic_data["desc"])
            qty = nic_data.get("qty", qty)
            nic = _get_or_create_nic(base_desc, nic_data["ports"],
                                     nic_data["speed"])
            db.session.add(ModelNicOption(
                model_id=model.id, nic_id=nic.id,
                quantity=qty, sort_order=i,
            ))

    db.session.commit()
    return jsonify({"message": f"Model '{model.name}' updated"})


# ── Delete model ─────────────────────────────────────────────────────────────

@admin_bp.route("/api/models/<int:model_id>", methods=["DELETE"])
def delete_model(model_id):
    model = db.get_or_404(Model, model_id)
    name = model.name
    db.session.delete(model)
    db.session.commit()
    return jsonify({"message": f"Model '{name}' deleted"})


# ── Export to Excel ──────────────────────────────────────────────────────────

# Full CPU catalog columns, in sheet order. The key is the ORM attribute; the
# value is the Excel header. Shared by the export and the importer so a column
# can only be added in one place — the old export carried just the first four
# and silently dropped every spec/benchmark field on a round-trip.
CPU_SHEET_COLUMNS = [
    ("description", "Description"),
    ("cores", "Cores"),
    ("threads", "Threads"),
    ("ghz", "GHz"),
    ("make", "Make"),
    ("family", "Family"),
    ("generation", "Generation"),
    ("model", "Model"),
    ("p_cores", "P-Cores"),
    ("e_cores", "E-Cores"),
    ("base_ghz", "Base GHz"),
    ("all_core_turbo_ghz", "All-Core Turbo GHz"),
    ("max_turbo_ghz", "Max Turbo GHz"),
    ("ecore_base_ghz", "E-Core Base GHz"),
    ("ecore_turbo_ghz", "E-Core Turbo GHz"),
    ("specrate_int", "SPECrate2017 int"),
    ("passmark_cpu_mark", "PassMark CPU Mark"),
    ("passmark_single", "PassMark Single"),
]
# Columns that must stay whole numbers when read back out of a spreadsheet.
_CPU_INT_COLUMNS = {"cores", "threads", "p_cores", "e_cores",
                    "passmark_cpu_mark", "passmark_single"}
_CPU_STR_COLUMNS = {"description", "make", "family", "generation", "model"}


@admin_bp.route("/api/export-models")
def export_models():
    """Download the whole admin-editable catalog as one .xlsx.

    This is the backup half of a backup/restore pair, so the sheet names and
    headers are exactly the ones _import_catalog_from_excel reads — re-importing
    this file has to reproduce the catalog it came from. It previously wrote its
    own sheet names ("CPU Options" where the importer looks for "Model CPU
    Options"), so only the Models sheet was ever read back and every model came
    in stripped of its CPU/RAM/storage/NIC options.

    Not covered here, by design: the switches and validated_* tables are seeded
    from models.py on every boot rather than admin-edited, so seed.py rebuilds
    them on any host.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="003A70")
    header_align = Alignment(horizontal="center", wrap_text=True)

    def sheet(name, headers, first=False):
        ws = wb.active if first else wb.create_sheet(name)
        ws.title = name
        ws.append(headers)
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
        return ws

    ws = sheet("Models", ["Name", "Status", "Category", "Form Factor", "Chassis",
                          "Socket", "PSU", "RAM Slots", "Min Nodes", "Cost",
                          "Validated Only", "Notes"], first=True)
    ws_cpu_cat = sheet("CPUs", [h for _, h in CPU_SHEET_COLUMNS])
    ws_nic_cat = sheet("NICs", ["Description", "Ports", "Speed"])
    ws_drv_cat = sheet("Drives", ["Type", "Size TB"])
    ws_cpu = sheet("Model CPU Options",
                   ["Model Name", "Qty", "Description", "Cores", "Threads", "GHz"])
    ws_ram = sheet("Model RAM Options", ["Model Name", "Size GB"])
    ws_stor = sheet("Model Storage", ["Model Name", "Type", "HDD Count", "SSD Count",
                                      "NVMe Count", "Drives Per Node"])
    ws_drv = sheet("Model Drive Options", ["Model Name", "Drive Type", "Size TB"])
    ws_nic = sheet("Model NIC Options",
                   ["Model Name", "Qty", "Description", "Ports", "Speed"])
    ws_iops = sheet("Drive IOPS", ["Drive Type", "IOPS"])
    ws_set = sheet("Sizing Settings", ["Key", "Value"])

    # Full catalog tables first — these carry the entries that no model
    # references yet, which a model-driven walk would never reach.
    for c in CpuCatalog.query.order_by(CpuCatalog.description).all():
        ws_cpu_cat.append([getattr(c, attr) for attr, _ in CPU_SHEET_COLUMNS])
    for n in NicCatalog.query.order_by(NicCatalog.description).all():
        ws_nic_cat.append([n.description, n.ports, n.speed])
    for d in DriveCatalog.query.order_by(DriveCatalog.drive_type,
                                         DriveCatalog.size_tb).all():
        ws_drv_cat.append([d.drive_type, d.size_tb])

    models = _model_query().order_by(Model.category, Model.name).all()
    for m in models:
        ws.append([m.name, m.status, m.category, m.form_factor, m.chassis,
                   m.socket, m.psu, m.ram_slots, m.min_nodes, m.cost_tier,
                   "Yes" if m.validated_only else "No", m.notes])

        for link in sorted(m.cpu_links, key=lambda l: l.sort_order):
            ws_cpu.append([m.name, link.quantity, link.cpu.description,
                           link.cpu.cores, link.cpu.threads, link.cpu.ghz])

        for ram in sorted(m.ram_options, key=lambda r: r.size_gb):
            ws_ram.append([m.name, ram.size_gb])

        sc = m.storage_config
        if sc:
            ws_stor.append([m.name, sc.storage_type, sc.hdd_count, sc.ssd_count,
                            sc.nvme_count, sc.drives_per_node])
            for link in sc.drive_links:
                ws_drv.append([m.name, link.drive.drive_type, link.drive.size_tb])

        for link in sorted(m.nic_links, key=lambda l: l.sort_order):
            ws_nic.append([m.name, link.quantity, link.nic.description,
                           link.nic.ports, link.nic.speed])

    # Admin-tuned sizing config travels with the catalog it applies to.
    for row in DriveTypeIops.query.order_by(DriveTypeIops.drive_type).all():
        ws_iops.append([row.drive_type, row.iops])
    for row in SizingSetting.query.order_by(SizingSetting.key).all():
        ws_set.append([row.key, row.value])

    for sh in wb.worksheets:
        for col in sh.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            sh.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="SC_Catalog_Backup.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument."
                              "spreadsheetml.sheet")



# ── Import from Excel ────────────────────────────────────────────────────────

@admin_bp.route("/api/import-models", methods=["POST"])
def import_models():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    if not f.filename or not f.filename.endswith(".xlsx"):
        return jsonify({"error": "File must be .xlsx"}), 400

    mode = request.form.get("mode", "add")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        f.save(tmp.name)
        tmp.close()
        result = _import_from_excel(tmp.name, mode)
        return jsonify(result)
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Model import failed: %s", e)
        return jsonify({"error": "Import failed. Check the file matches the template format."}), 400
    finally:
        os.unlink(tmp.name)


# ── Catalog import (CPUs / NICs / Drives in one file) ───────────────────────

@admin_bp.route("/api/import-catalog", methods=["POST"])
def import_catalog():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    if not f.filename or not f.filename.endswith(".xlsx"):
        return jsonify({"error": "File must be .xlsx"}), 400

    # "add" (default) never changes an existing row; "replace" makes the file
    # authoritative, which is what restoring a backup has to mean.
    mode = request.form.get("mode", "add")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        f.save(tmp.name)
        tmp.close()
        result = _import_catalog_from_excel(tmp.name, mode)
        return jsonify(result)
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Catalog import failed: %s", e)
        return jsonify({"error": "Import failed. Check the file matches the template format."}), 400
    finally:
        os.unlink(tmp.name)


def _model_option_maps(cpu_rows, ram_rows, stor_rows, drv_rows, nic_rows):
    """Group the per-model option rows (CPU / RAM / Storage / Drive / NIC) by
    model name. Shared by both Excel importers — they differ only in which sheet
    names these rows are read from, not in how the rows are shaped."""
    cpus = {}
    for r in cpu_rows:
        name = str(r.get("Model Name", "")).strip()
        if name:
            cpus.setdefault(name, []).append({
                "desc": str(r.get("Description", "")),
                "qty": int(r.get("Qty", 1) or 1),
                "cores": int(r.get("Cores", 0) or 0),
                "threads": int(r.get("Threads", 0) or 0),
                "ghz": float(r.get("GHz", 0) or 0),
            })

    ram = {}
    for r in ram_rows:
        name = str(r.get("Model Name", "")).strip()
        if name:
            ram.setdefault(name, []).append(int(r.get("Size GB", 0) or 0))

    stor = {}
    for r in stor_rows:
        name = str(r.get("Model Name", "")).strip()
        if name:
            stor[name] = {
                "type": str(r.get("Type", "nvme_only")).strip(),
                "hdd_count": int(r.get("HDD Count", 0) or 0) or None,
                "ssd_count": int(r.get("SSD Count", 0) or 0) or None,
                "nvme_count": int(r.get("NVMe Count", 0) or 0) or None,
                "drives_per_node": int(r.get("Drives Per Node", 0) or 0) or None,
            }

    drives = {}
    for r in drv_rows:
        name = str(r.get("Model Name", "")).strip()
        if name:
            drives.setdefault(name, []).append({
                "type": str(r.get("Drive Type", "")).strip(),
                "size_tb": float(r.get("Size TB", 0) or 0),
            })

    nics = {}
    for r in nic_rows:
        name = str(r.get("Model Name", "")).strip()
        if name:
            nics.setdefault(name, []).append({
                "desc": str(r.get("Description", "")),
                "qty": int(r.get("Qty", 1) or 1),
                "ports": int(r.get("Ports", 0) or 0),
                "speed": str(r.get("Speed", "")),
            })

    return cpus, ram, stor, drives, nics


def _cpu_fields_from_row(r):
    """Pull the CPU catalog columns out of one spreadsheet row, coercing each to
    its column type. Absent headers come back as None so the caller can tell
    "not in the file" from "explicitly blank"."""
    out = {}
    for attr, header in CPU_SHEET_COLUMNS:
        if header not in r:
            continue
        v = r[header]
        if attr in _CPU_STR_COLUMNS:
            v = str(v).strip() if v not in (None, "") else None
        elif attr in _CPU_INT_COLUMNS:
            v = _cpu_int(v)
        else:
            v = _cpu_num(v)
        out[attr] = v
    return out


def _apply_cpu_row(cpu, fields, overwrite):
    """Copy spreadsheet CPU fields onto a catalog row. Returns True if anything
    changed. With overwrite=False only currently-empty columns are filled, so an
    "add" import can enrich a sparse catalog without reverting later edits."""
    changed = False
    for attr, value in fields.items():
        if value is None or attr == "description":
            continue        # description is the identity key, never rewritten
        # cores/threads/ghz are non-nullable, so this skips them in "add" mode
        # and lets a restore correct them in "replace" mode.
        if not overwrite and getattr(cpu, attr) is not None:
            continue
        if getattr(cpu, attr) != value:
            setattr(cpu, attr, value)
            changed = True
    return changed


def _import_catalog_from_excel(file_path, mode="add"):
    """Restore a catalog workbook produced by export_models() (or hand-built
    from the template).

    mode="add"      — existing models are skipped; existing catalog rows keep
                      their values but have empty spec columns filled in.
    mode="replace"  — the file wins: models are rebuilt and catalog rows updated.
    """
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)

    def rows(*names):
        """First sheet that exists, by preference order. The second name is the
        pre-backup export's sheet name, kept so files exported by older builds
        still import — those wrote "CPU Options" where the template says
        "Model CPU Options"."""
        for n in names:
            r = _sheet_rows(wb, n)
            if r:
                return r
        return []

    overwrite = (mode == "replace")

    cpu_rows = rows("CPUs")
    nic_rows = rows("NICs")
    drive_rows = rows("Drives")
    model_rows = rows("Models")
    model_cpu_rows = rows("Model CPU Options", "CPU Options")
    model_ram_rows = rows("Model RAM Options", "RAM Options")
    model_stor_rows = rows("Model Storage", "Storage")
    model_drv_rows = rows("Model Drive Options", "Drive Options")
    model_nic_rows = rows("Model NIC Options", "NIC Options")
    iops_rows = rows("Drive IOPS")
    setting_rows = rows("Sizing Settings")
    wb.close()

    has_catalog = cpu_rows or nic_rows or drive_rows
    has_models = bool(model_rows)
    has_config = iops_rows or setting_rows

    if not has_catalog and not has_models and not has_config:
        return {"error": "No recognized sheets found. Expected: CPUs, NICs, "
                         "Drives, Models"}

    parts = []

    cpus_added = cpus_skipped = cpus_updated = 0
    for r in cpu_rows:
        fields = _cpu_fields_from_row(r)
        desc = fields.get("description")
        if not desc:
            continue
        existing = CpuCatalog.query.filter_by(description=desc).first()
        if existing:
            if _apply_cpu_row(existing, fields, overwrite):
                cpus_updated += 1
            else:
                cpus_skipped += 1
            continue
        cpu = CpuCatalog(
            description=desc,
            cores=fields.get("cores") or 0,
            threads=fields.get("threads") or 0,
            ghz=fields.get("ghz") or 0,
        )
        _apply_cpu_row(cpu, fields, True)
        db.session.add(cpu)
        cpus_added += 1
    if cpu_rows:
        parts.append(f"CPUs: {cpus_added} added, {cpus_updated} updated, "
                     f"{cpus_skipped} unchanged")

    nics_added = nics_skipped = nics_updated = 0
    for r in nic_rows:
        desc = str(r.get("Description", "")).strip()
        if not desc:
            continue
        ports = int(r.get("Ports", 0) or 0)
        speed = str(r.get("Speed", "")).strip()
        existing = NicCatalog.query.filter_by(description=desc).first()
        if existing:
            if overwrite and (existing.ports != ports or existing.speed != speed):
                existing.ports, existing.speed = ports, speed
                nics_updated += 1
            else:
                nics_skipped += 1
            continue
        db.session.add(NicCatalog(description=desc, ports=ports, speed=speed))
        nics_added += 1
    if nic_rows:
        parts.append(f"NICs: {nics_added} added, {nics_updated} updated, "
                     f"{nics_skipped} unchanged")

    drives_added = drives_skipped = 0
    for r in drive_rows:
        dtype = str(r.get("Type", "")).strip()
        size = float(r.get("Size TB", 0) or 0)
        if not dtype or size <= 0:
            continue
        if DriveCatalog.query.filter_by(drive_type=dtype, size_tb=size).first():
            drives_skipped += 1
            continue
        db.session.add(DriveCatalog(drive_type=dtype, size_tb=size))
        drives_added += 1
    if drive_rows:
        parts.append(f"Drives: {drives_added} added, {drives_skipped} unchanged")

    db.session.flush()

    models_created = models_replaced = models_skipped = 0
    if model_rows:
        cpus_by_model, ram_by_model, stor_by_model, mdrives_by_model, nics_by_model = \
            _model_option_maps(model_cpu_rows, model_ram_rows, model_stor_rows,
                               model_drv_rows, model_nic_rows)

        for r in model_rows:
            name = str(r.get("Name", "")).strip()
            if not name:
                continue
            existing = Model.query.filter_by(name=name).first()
            if existing and not overwrite:
                models_skipped += 1
                continue

            storage_data = stor_by_model.get(name, {})
            for drv in mdrives_by_model.get(name, []):
                key = f"{drv['type'].lower()}_options_tb"
                storage_data.setdefault(key, []).append(drv["size_tb"])

            model_data = {
                "name": name,
                "status": str(r.get("Status", "Active")).strip(),
                "category": str(r.get("Category", "")).strip(),
                "form_factor": str(r.get("Form Factor", "") or "").strip() or None,
                "chassis": str(r.get("Chassis", "") or "").strip() or None,
                "socket": str(r.get("Socket", "single") or "single").strip(),
                "psu": str(r.get("PSU", "") or "").strip() or None,
                "ram_slots": int(r.get("RAM Slots", 0) or 0),
                "min_nodes": int(r.get("Min Nodes", 1) or 1),
                "cost_tier": float(r["Cost"]) if r.get("Cost") not in (None, "") else 5.0,
                "validated_only": str(r.get("Validated Only", "")).strip().lower()
                                  in ("yes", "true", "1"),
                "notes": str(r.get("Notes", "") or "").strip() or None,
                "cpu_options": cpus_by_model.get(name, []),
                "ram_options_gb": ram_by_model.get(name, []),
                "storage": storage_data,
                "nic_options": nics_by_model.get(name, []),
            }
            if existing:
                db.session.delete(existing)
                db.session.flush()
                _build_model(model_data)
                models_replaced += 1
            else:
                _build_model(model_data)
                models_created += 1

        parts.append(f"Models: {models_created} created, {models_replaced} replaced, "
                     f"{models_skipped} skipped (existing)")

    # Admin-tuned sizing config. Both tables are pure key/value, so a restore is
    # an upsert; in "add" mode an existing key is left as the host has it.
    iops_set = 0
    for r in iops_rows:
        dtype = str(r.get("Drive Type", "")).strip()
        val = _cpu_int(r.get("IOPS"))
        if not dtype or val is None:
            continue
        row = DriveTypeIops.query.filter_by(drive_type=dtype).first()
        if row is None:
            db.session.add(DriveTypeIops(drive_type=dtype, iops=val))
            iops_set += 1
        elif overwrite and row.iops != val:
            row.iops = val
            iops_set += 1
    if iops_rows:
        parts.append(f"Drive IOPS: {iops_set} set")

    settings_set = 0
    for r in setting_rows:
        key = str(r.get("Key", "")).strip()
        val = _cpu_num(r.get("Value"))
        if not key or val is None:
            continue
        row = SizingSetting.query.filter_by(key=key).first()
        if row is None:
            db.session.add(SizingSetting(key=key, value=val))
            settings_set += 1
        elif overwrite and row.value != val:
            row.value = val
            settings_set += 1
    if setting_rows:
        parts.append(f"Sizing settings: {settings_set} set")

    db.session.commit()

    return {"message": ". ".join(parts)}



# ── Catalog template download ──────────────────────────────────────────────

@admin_bp.route("/api/catalog-template")
def catalog_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="003A70")
    header_align = Alignment(horizontal="center", wrap_text=True)
    example_font = Font(italic=True, color="888888")

    def style_headers(ws):
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align

    def example_rows(ws, rows):
        for row in rows:
            ws.append(row)
        for r in range(2, 2 + len(rows)):
            for c in ws[r]:
                c.font = example_font

    ex = "HC9999F"

    ws_cpu = wb.active
    ws_cpu.title = "CPUs"
    # Same column set the export writes, so a template-authored file and a
    # backup are the same format. Everything past GHz is optional — but
    # SPECrate2017 int (or PassMark CPU Mark) is what perf-based sizing reads.
    ws_cpu.append([h for _, h in CPU_SHEET_COLUMNS])
    style_headers(ws_cpu)
    example_rows(ws_cpu, [
        ["Xeon Gold 6526Y 16C/32T 3.5GHz", 16, 32, 3.5, "Intel", "Xeon Gold",
         "Emerald Rapids", "6526Y", 16, 0, 2.8, 3.5, 4.0, None, None,
         143.0, None, None],
        ["Silver 4516Y+ 24C/48T 2.9GHz", 24, 48, 2.9, "Intel", "Xeon Silver",
         "Emerald Rapids", "4516Y+", 24, 0, 2.2, 2.9, 3.4, None, None,
         120.0, None, None],
    ])

    ws_nic = wb.create_sheet("NICs")
    ws_nic.append(["Description", "Ports", "Speed"])
    style_headers(ws_nic)
    example_rows(ws_nic, [
        ["10GbE SFP+ 4-port Network Card (Intel X710)", 4, "10GbE"],
        ["25GbE SFP28 2-port OCP Network Card (Intel E810)", 2, "25GbE"],
    ])

    ws_drv = wb.create_sheet("Drives")
    ws_drv.append(["Type", "Size TB"])
    style_headers(ws_drv)
    example_rows(ws_drv, [
        ["NVMe", 3.84],
        ["SSD", 1.92],
    ])

    ws_mod = wb.create_sheet("Models")
    ws_mod.append(["Name", "Status", "Category", "Form Factor", "Chassis",
                   "Socket", "PSU", "RAM Slots", "Min Nodes", "Cost",
                   "Validated Only", "Notes"])
    style_headers(ws_mod)
    example_rows(ws_mod, [
        [ex, "Active", "1U All-Flash", "1U Rack", "Dell PowerEdge R660",
         "single", "2x 800W", 16, 3, 28, "No", None],
    ])

    ws_mcpu = wb.create_sheet("Model CPU Options")
    ws_mcpu.append(["Model Name", "Qty", "Description", "Cores", "Threads", "GHz"])
    style_headers(ws_mcpu)
    example_rows(ws_mcpu, [
        [ex, 1, "Xeon Gold 6526Y 16C/32T 3.5GHz", 16, 32, 3.5],
        [ex, 1, "Silver 4516Y+ 24C/48T 2.9GHz", 24, 48, 2.9],
    ])

    ws_mram = wb.create_sheet("Model RAM Options")
    ws_mram.append(["Model Name", "Size GB"])
    style_headers(ws_mram)
    example_rows(ws_mram, [
        [ex, 64], [ex, 128], [ex, 256],
    ])

    ws_mstor = wb.create_sheet("Model Storage")
    ws_mstor.append(["Model Name", "Type", "HDD Count", "SSD Count",
                     "NVMe Count", "Drives Per Node"])
    style_headers(ws_mstor)
    example_rows(ws_mstor, [
        [ex, "nvme_only", None, None, None, 10],
    ])

    ws_mdrv = wb.create_sheet("Model Drive Options")
    ws_mdrv.append(["Model Name", "Drive Type", "Size TB"])
    style_headers(ws_mdrv)
    example_rows(ws_mdrv, [
        [ex, "NVMe", 3.84],
        [ex, "NVMe", 7.68],
    ])

    ws_mnic = wb.create_sheet("Model NIC Options")
    ws_mnic.append(["Model Name", "Qty", "Description", "Ports", "Speed"])
    style_headers(ws_mnic)
    example_rows(ws_mnic, [
        [ex, 1, "10GbE SFP+ 4-port Network Card (Intel X710)", 4, "10GbE"],
        [ex, 1, "25GbE SFP28 2-port OCP Network Card (Intel E810)", 2, "25GbE"],
    ])

    ws_iops = wb.create_sheet("Drive IOPS")
    ws_iops.append(["Drive Type", "IOPS"])
    style_headers(ws_iops)
    example_rows(ws_iops, [["NVMe", 75000], ["SSD", 20000], ["HDD", 150]])

    ws_set = wb.create_sheet("Sizing Settings")
    ws_set.append(["Key", "Value"])
    style_headers(ws_set)
    example_rows(ws_set, [["iops_read_fraction", 0.70],
                          ["iops_replication_factor", 2]])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="SC_Import_Template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Helpers ──────────────────────────────────────────────────────────────────

def _get_or_create_cpu(desc, cores, threads, ghz):
    cpu = CpuCatalog.query.filter_by(description=desc).first()
    if not cpu:
        cpu = CpuCatalog(description=desc, cores=cores, threads=threads, ghz=ghz)
        db.session.add(cpu)
        db.session.flush()
    return cpu


def _get_or_create_nic(desc, ports, speed):
    nic = NicCatalog.query.filter_by(description=desc).first()
    if not nic:
        nic = NicCatalog(description=desc, ports=ports, speed=speed)
        db.session.add(nic)
        db.session.flush()
    return nic


def _get_or_create_drive(drive_type, size_tb):
    drive = DriveCatalog.query.filter_by(drive_type=drive_type, size_tb=size_tb).first()
    if not drive:
        drive = DriveCatalog(drive_type=drive_type, size_tb=size_tb)
        db.session.add(drive)
        db.session.flush()
    return drive


def _build_model(data):
    model = Model(
        name=data["name"],
        status=data.get("status", "Active"),
        category=data.get("category", ""),
        form_factor=data.get("form_factor"),
        chassis=data.get("chassis"),
        socket=data.get("socket", "single"),
        psu=data.get("psu"),
        ram_slots=data.get("ram_slots", 0),
        min_nodes=data.get("min_nodes", 1),
        cost_tier=float(data["cost_tier"]) if data.get("cost_tier") not in (None, "") else 5.0,
        validated_only=bool(data.get("validated_only", False)),
        notes=data.get("notes"),
    )
    db.session.add(model)
    db.session.flush()

    for i, cpu_data in enumerate(data.get("cpu_options", [])):
        qty, base_desc = _parse_quantity(cpu_data["desc"])
        qty = cpu_data.get("qty", qty)
        cpu = _get_or_create_cpu(base_desc, cpu_data["cores"],
                                 cpu_data["threads"], cpu_data["ghz"])
        db.session.add(ModelCpuOption(
            model_id=model.id, cpu_id=cpu.id,
            quantity=qty, sort_order=i,
        ))

    for size in data.get("ram_options_gb", []):
        db.session.add(RamOption(model_id=model.id, size_gb=size))

    if "storage" in data:
        _add_storage(model, data["storage"])

    for i, nic_data in enumerate(data.get("nic_options", [])):
        qty, base_desc = _parse_quantity(nic_data["desc"])
        qty = nic_data.get("qty", qty)
        nic = _get_or_create_nic(base_desc, nic_data["ports"],
                                 nic_data["speed"])
        db.session.add(ModelNicOption(
            model_id=model.id, nic_id=nic.id,
            quantity=qty, sort_order=i,
        ))

    return model


def _add_storage(model, storage):
    sc = StorageConfig(
        model_id=model.id,
        storage_type=storage.get("type", "nvme_only"),
        hdd_count=storage.get("hdd_count"),
        ssd_count=storage.get("ssd_count"),
        nvme_count=storage.get("nvme_count"),
        drives_per_node=storage.get("drives_per_node"),
    )
    if storage.get("type") == "cloud" and "options" in storage:
        sc.cloud_tiers = "|".join(storage["options"])
    db.session.add(sc)
    db.session.flush()

    for dtype_key, dtype_label in [
        ("hdd_options_tb", "HDD"),
        ("ssd_options_tb", "SSD"),
        ("nvme_options_tb", "NVMe"),
    ]:
        for size in storage.get(dtype_key, []):
            drive = _get_or_create_drive(dtype_label, size)
            db.session.add(StorageConfigDrive(
                storage_config_id=sc.id, drive_id=drive.id,
            ))


def _import_from_excel(file_path, mode):
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)

    model_rows = _sheet_rows(wb, "Models")
    cpu_rows = _sheet_rows(wb, "CPU Options")
    ram_rows = _sheet_rows(wb, "RAM Options")
    stor_rows = _sheet_rows(wb, "Storage")
    drv_rows = _sheet_rows(wb, "Drive Options")
    nic_rows = _sheet_rows(wb, "NIC Options")
    wb.close()

    if not model_rows:
        return {"error": "No models found in the 'Models' sheet"}

    cpus_by_model, ram_by_model, stor_by_model, drives_by_model, nics_by_model = \
        _model_option_maps(cpu_rows, ram_rows, stor_rows, drv_rows, nic_rows)

    created = 0
    updated = 0
    skipped = []

    for r in model_rows:
        name = str(r.get("Name", "")).strip()
        if not name:
            continue

        existing = Model.query.filter_by(name=name).first()

        if existing and mode == "add":
            skipped.append(name)
            continue

        storage_data = stor_by_model.get(name, {})
        drives = drives_by_model.get(name, [])
        for drv in drives:
            key = f"{drv['type'].lower()}_options_tb"
            storage_data.setdefault(key, []).append(drv["size_tb"])

        model_data = {
            "name": name,
            "status": str(r.get("Status", "Active")).strip(),
            "category": str(r.get("Category", "")).strip(),
            "form_factor": str(r.get("Form Factor", "") or "").strip() or None,
            "chassis": str(r.get("Chassis", "") or "").strip() or None,
            "socket": str(r.get("Socket", "single") or "single").strip(),
            "psu": str(r.get("PSU", "") or "").strip() or None,
            "ram_slots": int(r.get("RAM Slots", 0) or 0),
            "min_nodes": int(r.get("Min Nodes", 1) or 1),
            # Both are written by export_models(); reading them back was missed,
            # so every imported model silently fell back to cost_tier 5.0 and
            # validated_only False — and cost_tier feeds the ranker.
            "cost_tier": float(r["Cost"]) if r.get("Cost") not in (None, "") else 5.0,
            "validated_only": str(r.get("Validated Only", "")).strip().lower()
                              in ("yes", "true", "1"),
            "notes": str(r.get("Notes", "") or "").strip() or None,
            "cpu_options": cpus_by_model.get(name, []),
            "ram_options_gb": ram_by_model.get(name, []),
            "storage": storage_data,
            "nic_options": nics_by_model.get(name, []),
        }

        if existing and mode == "replace":
            db.session.delete(existing)
            db.session.flush()
            _build_model(model_data)
            updated += 1
        else:
            _build_model(model_data)
            created += 1

    db.session.commit()

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total_in_file": len(model_rows),
    }
